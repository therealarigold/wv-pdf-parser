import os, sys, json, re, io, asyncio

from http.server import HTTPServer, BaseHTTPRequestHandler
import subprocess, sys

def ensure_chromium():
    try:
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True, text=True, timeout=300
        )
        print("Chromium ready:", result.returncode)
    except Exception as e:
        print(f"Chromium install warning: {e}")

import json, io, re, pdfplumber, os, urllib.request, urllib.parse
from datetime import datetime, timezone
from html.parser import HTMLParser

WV_COUNTIES = {
    "BARBOUR","BERKELEY","BOONE","BRAXTON","BROOKE","CABELL","CALHOUN","CLAY",
    "DODDRIDGE","FAYETTE","GILMER","GRANT","GREENBRIER","HAMPSHIRE","HANCOCK",
    "HARDY","HARRISON","JACKSON","JEFFERSON","KANAWHA","LEWIS","LINCOLN","LOGAN",
    "MARION","MARSHALL","MASON","MCDOWELL","MERCER","MINERAL","MINGO","MONONGALIA",
    "MONROE","MORGAN","NICHOLAS","OHIO","PENDLETON","PLEASANTS","POCAHONTAS",
    "PRESTON","PUTNAM","RALEIGH","RANDOLPH","RITCHIE","ROANE","SUMMERS","TAYLOR",
    "TUCKER","TYLER","UPSHUR","WAYNE","WEBSTER","WETZEL","WIRT","WOOD","WYOMING"
}

# ── ALL 55 COUNTY CAMA REGISTRY ───────────────────────────────────────────────
# 54 counties use wvassessor.com (GST platform) with prc.aspx?PARID=
# Wood County uses its own URL at inquiries.woodcountywv.com
# PARID format: DD++MMMMPPPP0000000 (district 2dig, map 4dig, parcel 4dig, 7 zeros)

def build_parid_variants(dist, map_num, parcel):
    """Return list of PARID variants to try - different counties use different formats."""
    dist_int = int(dist) if str(dist).isdigit() else 1
    dist2 = str(dist_int).zfill(2)
    map_str = str(map_num).zfill(4)
    parcel_str = str(parcel).zfill(4)
    zeros = "0000000"
    return [
        f"{dist2}++{map_str}{parcel_str}{zeros}",   # Format A: Wood, Marion, Kanawha
        f"{dist2}+++{map_str}{parcel_str}{zeros}",  # Format B: Wyoming, Hampshire, Preston
        f"{dist2}+{map_str}{parcel_str}{zeros}",    # Format C: single plus
        f"{dist2}  {map_str}{parcel_str}{zeros}",   # Format D: spaces
    ]

def build_standard_parid(dist, map_num, parcel):
    return build_parid_variants(dist, map_num, parcel)[0]

def build_wood_parid(dist, map_num, parcel):
    return build_parid_variants(dist, map_num, parcel)[0]

# Standard counties — all use COUNTY.wvassessor.com
STANDARD_CAMA_COUNTIES = [
    "BARBOUR","BERKELEY","BOONE","BRAXTON","BROOKE","CABELL","CALHOUN","CLAY",
    "DODDRIDGE","FAYETTE","GILMER","GRANT","GREENBRIER","HAMPSHIRE","HANCOCK",
    "HARDY","HARRISON","JACKSON","JEFFERSON","KANAWHA","LEWIS","LINCOLN","LOGAN",
    "MARION","MARSHALL","MASON","MCDOWELL","MERCER","MINERAL","MINGO","MONONGALIA",
    "MONROE","MORGAN","NICHOLAS","OHIO","PENDLETON","PLEASANTS","POCAHONTAS",
    "PRESTON","PUTNAM","RALEIGH","RANDOLPH","RITCHIE","ROANE","SUMMERS","TAYLOR",
    "TUCKER","TYLER","UPSHUR","WAYNE","WEBSTER","WETZEL","WIRT","WYOMING"
]

# IDX document search counties
IDX_COUNTIES = {
    "BARBOUR": "http://129.71.117.241/WEBInquiry/Default.aspx",
    "BROOKE":  "http://129.71.117.252/",
    "CABELL":  "http://www.recordscabellcountyclerk.org/Default.aspx",
    "DODDRIDGE":"http://129.71.205.241/",
    "FAYETTE": "http://129.71.202.7/",
    "GILMER":  "http://www.gilmercountywv.gov/idxsearch/",
    "GRANT":   "http://129.71.112.124/",
    "GREENBRIER":"http://129.71.205.208/",
    "HAMPSHIRE":"http://129.71.205.207/idxsearch",
    "HANCOCK": "https://hancockwv.compiled-technologies.com/",
    "HARRISON":"http://lookup.harrisoncountywv.com/",
    "JEFFERSON":"http://documents.jeffersoncountywv.org/",
    "LEWIS":   "http://inquiry.lewiscountywv.org/",
    "LINCOLN": "http://129.71.206.62/Default.aspx",
    "MARSHALL":"http://129.71.117.225/",
    "OHIO":    "https://ohiocountywvclerk.com/",
    "WETZEL":  "https://www.wetzelcountywv.gov/county-clerk-responsibilities",
    "WIRT":    "http://records.wirtcountywv.net/",
    "WOOD":    "https://inquiries.woodcountywv.com/legacywebinquiry/default.aspx",
}

def get_cama_url(county):
    if county == "WOOD":
        return "https://inquiries.woodcountywv.com/CAMA/prc.aspx?PARID={parid}"
    return f"https://{county.lower()}.wvassessor.com/prc.aspx?PARID={{parid}}"

def get_county_registry():
    result = {}
    for c in STANDARD_CAMA_COUNTIES:
        result[c] = {"has_cama": True, "has_idx": c in IDX_COUNTIES,
                     "cama_url": f"https://{c.lower()}.wvassessor.com/",
                     "idx_url": IDX_COUNTIES.get(c, "")}
    result["WOOD"] = {"has_cama": True, "has_idx": True,
                      "cama_url": "https://inquiries.woodcountywv.com/CAMA/",
                      "idx_url": IDX_COUNTIES["WOOD"]}
    return result

# ── HTML PARSERS ──────────────────────────────────────────────────────────────

class CAMAParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.data = {"owner":"","mailing_address":"","deed_book":"","deed_page":"",
                     "sales_history":[],"assessments":[],"legal_description":"",
                     "district":"","map_parcel":"","raw_text":""}
        self._text_parts = []
        self._current_row = []
        self._current_cell = ""
        self._table_rows = []
        self._current_tables = []
        self._cell_depth = 0

    def handle_starttag(self, tag, attrs):
        if tag == "table": self._current_tables.append([])
        elif tag == "tr": self._current_row = []
        elif tag in ("td","th"): self._current_cell = ""; self._cell_depth += 1

    def handle_endtag(self, tag):
        if tag == "table" and self._current_tables:
            self._table_rows.append(self._current_tables.pop())
        elif tag == "tr" and self._current_tables:
            self._current_tables[-1].append(self._current_row[:])
        elif tag in ("td","th"):
            self._cell_depth -= 1
            self._current_row.append(self._current_cell.strip())

    def handle_data(self, data):
        d = data.strip()
        if d:
            self._text_parts.append(d)
            if self._cell_depth > 0: self._current_cell += " " + d

    def extract(self):
        full_text = " ".join(self._text_parts)
        self.data["raw_text"] = full_text

        m = re.search(r'CURRENT OWNER:\s*([A-Z][^\n]+?)(?:MAILING ADDRESS:|DEED BOOK)', full_text, re.I)
        if m: self.data["owner"] = m.group(1).strip()

        m = re.search(r'MAILING ADDRESS:\s*([^\n]+?)(?:DEED BOOK|SALES HISTORY)', full_text, re.I)
        if m: self.data["mailing_address"] = m.group(1).strip()

        m = re.search(r'DEED BOOK/PAGE:\s*(\d+)/(\d+)', full_text, re.I)
        if m: self.data["deed_book"] = m.group(1); self.data["deed_page"] = m.group(2)

        m = re.search(r'LEGAL DESCRIPTION:\s*([^\n]+?)(?:CURRENT OWNER|TAXING|$)', full_text, re.I)
        if m: self.data["legal_description"] = m.group(1).strip()

        m = re.search(r'TAXING DISTRICT:\s*([^\n]+?)(?:STREET|PARCEL|$)', full_text, re.I)
        if m: self.data["district"] = m.group(1).strip()

        # Also try alternate patterns for wvassessor.com layout
        if not self.data["deed_book"]:
            m = re.search(r'Book[:/\s]+(\d+)\s*[/\s,]+\s*Page[:/\s]+(\d+)', full_text, re.I)
            if m: self.data["deed_book"] = m.group(1); self.data["deed_page"] = m.group(2)

        if not self.data["owner"]:
            m = re.search(r'Owner[:\s]+([A-Z][A-Z\s,]+?)(?:Address|Deed|Mailing|$)', full_text, re.I)
            if m: self.data["owner"] = m.group(1).strip()

        for table in self._table_rows:
            for row in table:
                cells = [c.strip() for c in row if c.strip()]
                if len(cells) >= 4:
                    dm = re.match(r'(\d{1,2}/\d{1,2}/\d{4})', cells[0])
                    pm = re.match(r'\$[\d,]+', cells[1]) if len(cells) > 1 else None
                    if dm and pm:
                        try:
                            self.data["sales_history"].append({
                                "date": cells[0].split()[0],
                                "price": cells[1],
                                "book": cells[2] if len(cells) > 2 else "",
                                "page": cells[3] if len(cells) > 3 else "",
                            })
                        except: pass
                if len(cells) >= 4 and re.match(r'^20\d{2}$', cells[0]):
                    try:
                        self.data["assessments"].append({
                            "year": cells[0],
                            "land": cells[1] if len(cells) > 1 else "",
                            "building": cells[2] if len(cells) > 2 else "",
                            "total": cells[3] if len(cells) > 3 else "",
                            "assessed": cells[4] if len(cells) > 4 else "",
                        })
                    except: pass
        return self.data


class IDXParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.records = []
        self._in_table = False
        self._headers = []
        self._rows = []
        self._current_row = []
        self._current_cell = ""
        self._cell_depth = 0
        self._header_row = True

    def handle_starttag(self, tag, attrs):
        if tag == "table": self._in_table = True
        elif tag == "tr": self._current_row = []
        elif tag in ("td","th"): self._current_cell = ""; self._cell_depth += 1

    def handle_endtag(self, tag):
        if tag == "tr" and self._in_table:
            row = [c.strip() for c in self._current_row]
            if any(row):
                if self._header_row and any(h in " ".join(row).upper() for h in ["GRANTOR","GRANTEE","BOOK","DATE","TYPE","INSTRUMENT"]):
                    self._headers = row; self._header_row = False
                elif not self._header_row:
                    self._rows.append(row)
        elif tag in ("td","th"):
            self._cell_depth -= 1
            self._current_row.append(self._current_cell.strip())

    def handle_data(self, data):
        d = data.strip()
        if d and self._cell_depth > 0: self._current_cell += " " + d

    def extract(self):
        for row in self._rows:
            if len(row) < 3: continue
            record = {}
            for i, header in enumerate(self._headers):
                if i < len(row): record[header.lower().replace(" ","_")] = row[i]
            if record: self.records.append(record)
        return self.records


# ── FETCH ─────────────────────────────────────────────────────────────────────

def fetch_url(url, timeout=20, post_data=None):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.9",
    }
    req = urllib.request.Request(url, headers=headers)
    if post_data:
        req.data = urllib.parse.urlencode(post_data).encode()
        req.add_header("Content-Type","application/x-www-form-urlencoded")
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        charset = "utf-8"
        ct = resp.headers.get("Content-Type","")
        if "charset=" in ct: charset = ct.split("charset=")[1].strip().split(";")[0]
        return resp.read().decode(charset, errors="replace")


# ── CAMA LOOKUP ───────────────────────────────────────────────────────────────

def lookup_cama(county_name, dist, map_num, parcel):
    county = county_name.upper().replace(" COUNTY","").strip()
    has_cama = county in STANDARD_CAMA_COUNTIES or county == "WOOD"
    if not has_cama:
        return {"success": False, "error": f"No CAMA system for {county} County"}
    try:
        cama_url_tmpl = get_cama_url(county)
        # Try multiple PARID formats until we get real data
        parids = build_parid_variants(dist, map_num, parcel)
        html = None
        parid = parids[0]
        data = None
        for p in parids:
            try:
                test_url = cama_url_tmpl.format(parid=p)
                test_html = fetch_url(test_url, timeout=15)
                # Check if we got real data (not just an empty/error page)
                if 'CURRENT OWNER' in test_html.upper() or 'DEED BOOK' in test_html.upper() or 'OWNER NAME' in test_html.upper() or 'SALES HISTORY' in test_html.upper():
                    html = test_html
                    parid = p
                    url = test_url
                    break
            except:
                continue
        if not html:
            # Fall back to first format even if no data found
            parid = parids[0]
            url = cama_url_tmpl.format(parid=parid)
            html = fetch_url(url, timeout=15)
        parser = CAMAParser()
        parser.feed(html)
        data = parser.extract()

        years_owned = None
        acquisition_year = None
        if data["sales_history"]:
            sorted_sales = sorted(data["sales_history"], key=lambda s: s["date"], reverse=True)
            try:
                yr = int(sorted_sales[0]["date"].split("/")[-1])
                years_owned = datetime.now().year - yr
                acquisition_year = yr
            except: pass

        return {
            "success": True,
            "county": county,
            "parid": parid,
            "cama_url": url,
            "owner": data["owner"],
            "mailing_address": data["mailing_address"],
            "deed_book": data["deed_book"],
            "deed_page": data["deed_page"],
            "legal_description": data["legal_description"],
            "district_label": data["district"],
            "years_owned": years_owned,
            "acquisition_year": acquisition_year,
            "sales_history": data["sales_history"],
            "assessments": data["assessments"][:5],
        }
    except Exception as e:
        return {"success": False, "error": f"CAMA lookup failed: {str(e)}", "cama_url": url if 'url' in dir() else ""}


# ── IDX SEARCH ────────────────────────────────────────────────────────────────

def search_idx(county_name, grantor_name):
    county = county_name.upper().replace(" COUNTY","").strip()
    idx_url = IDX_COUNTIES.get(county)
    if not idx_url:
        return {"success": False, "error": f"No IDX system for {county} County", "county_has_idx": False}
    try:
        search_url = idx_url + f"?name={urllib.parse.quote(grantor_name)}&searchType=grantor"
        html = fetch_url(search_url, timeout=20)
        parser = IDXParser()
        parser.feed(html)
        records = parser.extract()
        return {
            "success": True, "county": county, "idx_url": idx_url,
            "grantor_searched": grantor_name,
            "records_found": len(records), "records": records[:25],
        }
    except Exception as e:
        return {"success": False, "error": f"IDX search failed: {str(e)}",
                "county_has_idx": True, "idx_url": idx_url}



# ── CLAUDE AI PROXY ───────────────────────────────────────────────────────────

def call_claude(prompt):
    """Call Anthropic Claude API server-side and return the analysis text."""
    if not prompt:
        return {"success": False, "error": "No prompt provided"}
    try:
        api_key = os.environ.get("ANTHROPIC_API_KEY", "")
        if not api_key:
            return {"success": False, "error": "ANTHROPIC_API_KEY not set on server"}

        payload = json.dumps({
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 1200,
            "messages": [{"role": "user", "content": prompt}]
        }).encode()

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            }
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
            text = result.get("content", [{}])[0].get("text", "")
            return {"success": True, "text": text}
    except Exception as e:
        return {"success": False, "error": f"Claude API error: {str(e)}"}


# ── MAPWV PARCEL LOOKUP ───────────────────────────────────────────────────────

def fetch_mapwv_owner(mapwv_url, county_key, map_num, parcel_num):
    """
    Fetch the current owner from MapWV by hitting their parcel data API.
    The pid is in the URL we already build correctly in the app.
    MapWV loads data via: /parcel/php/getparceldata.php?pid=XX-XX-XXXX-XXXX-XXXX
    """
    try:
        pid = None
        if 'pid=' in mapwv_url:
            pid = mapwv_url.split('pid=')[1].split('&')[0]

        if not pid:
            return {'success': False, 'error': 'No pid in URL'}

        # Try MapWV internal data endpoints
        endpoints = [
            f'https://mapwv.gov/parcel/php/getparceldata.php?pid={pid}',
            f'https://mapwv.gov/parcel/php/getparcelinfo.php?pid={pid}',
            f'https://mapwv.gov/parcel/php/parcelinfo.php?pid={pid}',
        ]

        for endpoint in endpoints:
            try:
                html = fetch_url(endpoint, timeout=12)
                if not html or len(html) < 10:
                    continue
                # Try JSON first
                try:
                    data = json.loads(html)
                    # Look for owner in various key names
                    for key in ['OwnerName','owner','OWNERNAME','Owner','fullownername','FullOwnerName']:
                        if key in data and data[key]:
                            addr = data.get('OwnerAddress','') or data.get('address','') or data.get('OWNERADDRESS','')
                            return {'success':True,'owner':str(data[key]).strip(),'address':str(addr).strip(),'pid':pid}
                    # If it's a list
                    if isinstance(data, list) and len(data) > 0:
                        item = data[0]
                        for key in ['OwnerName','owner','OWNERNAME','Owner']:
                            if key in item and item[key]:
                                return {'success':True,'owner':str(item[key]).strip(),'address':'','pid':pid}
                except (json.JSONDecodeError, TypeError):
                    # Try regex on raw HTML/text
                    m = re.search(r'[Oo]wner[^:]*:[\s"]*([A-Z][A-Z ,&.]+)', html)
                    if m:
                        return {'success':True,'owner':m.group(1).strip(),'address':'','pid':pid}
            except Exception:
                continue

        return {'success': False, 'error': 'MapWV API did not return owner data', 'pid': pid}

    except Exception as e:
        return {'success': False, 'error': str(e)}


# ── MAPWV ASSESSMENT DETAIL LOOKUP ───────────────────────────────────────────

def fetch_assessment_detail(map_pid):
    """
    Fetch MapWV Assessment Detail page and parse all fields.
    map_pid example: "22-01-0011-0010-0005"
    Assessment URL: https://mapwv.gov/Assessment/Detail/?PID=22010011001000050000
    """
    try:
        # Build assessment PID: remove dashes, pad to 20 chars with zeros
        clean = map_pid.replace('-', '')
        assessment_pid = clean.ljust(20, '0')
        url = f'https://mapwv.gov/Assessment/Detail/?PID={assessment_pid}'

        html = fetch_url(url, timeout=25)
        if not html or len(html) < 500:
            return {'success': False, 'error': 'Empty or invalid response from assessment page', 'assessment_url': url}

        result = {
            'success': True,
            'assessment_url': url,
            'pid': map_pid,
            'assessment_pid': assessment_pid,
            'owner': '',
            'mailing_address': '',
            'tax_class': '',
            'deed_book': '',
            'deed_page': '',
            'legal_description': '',
            'property_class': '',
            'land_use': '',
            'total_appraisal': '',
            'sales_history': [],
            'parcel_history': [],
        }

        # ── Strip scripts/styles, get clean text ──────────────────
        clean_html = re.sub(r'<script[^>]*>.*?</script>', ' ', html, flags=re.DOTALL|re.I)
        clean_html = re.sub(r'<style[^>]*>.*?</style>', ' ', clean_html, flags=re.DOTALL|re.I)

        # ── Owner name ─────────────────────────────────────────────
        # The page has: <td>Owner(s)</td><td>KEENEY DON</td>
        m = re.search(r'Owner[(]s[)]\s*</td>\s*<td[^>]*>\s*([^<]+)', clean_html, re.I)
        if not m:
            m = re.search(r'Owner[(]s[)][^<]*<[^>]+>\s*([A-Z][A-Z\s,&.\-]+)', clean_html, re.I)
        if m:
            result['owner'] = m.group(1).strip()

        # ── Mailing address ────────────────────────────────────────
        m = re.search(r'Mailing\s*Address\s*</td>\s*<td[^>]*>\s*([^<]+)', clean_html, re.I)
        if not m:
            m = re.search(r'Mailing\s*Address[^<]*<[^>]+>\s*([A-Z0-9][^<]{5,})', clean_html, re.I)
        if m:
            result['mailing_address'] = m.group(1).strip()

        # ── Tax class ──────────────────────────────────────────────
        # "Tax Class" header then value in next cell
        m = re.search(r'Tax\s*Class\s*</th>.*?<td[^>]*>\s*(\d+)', clean_html, re.I|re.DOTALL)
        if not m:
            m = re.search(r'Tax\s*Class[^<]*</td>\s*<td[^>]*>\s*(\d+)', clean_html, re.I)
        if m:
            result['tax_class'] = m.group(1).strip()

        # ── Book / Page ────────────────────────────────────────────
        m = re.search(r'Book\s*/\s*Page\s*</th>.*?<td[^>]*>\s*(\d+)\s*/\s*(\d+)', clean_html, re.I|re.DOTALL)
        if not m:
            m = re.search(r'>(\d{3,})\s*/\s*(\d{2,})<', clean_html)
        if m:
            result['deed_book'] = m.group(1).strip()
            result['deed_page'] = m.group(2).strip()

        # ── Legal description ──────────────────────────────────────
        m = re.search(r'Legal\s*Description\s*</th>.*?<td[^>]*>\s*([^<]{5,})', clean_html, re.I|re.DOTALL)
        if not m:
            m = re.search(r'Legal\s*Description[^<]*</td>\s*<td[^>]*>([^<]{5,})', clean_html, re.I)
        if m:
            result['legal_description'] = m.group(1).strip()

        # ── Property class ─────────────────────────────────────────
        m = re.search(r'Property\s*Class\s*</td>\s*<td[^>]*>\s*([^<]+)', clean_html, re.I)
        if m:
            result['property_class'] = m.group(1).strip()

        # ── Land use ───────────────────────────────────────────────
        m = re.search(r'Land\s*Use\s*</td>\s*<td[^>]*>\s*([^<]+)', clean_html, re.I)
        if m:
            result['land_use'] = m.group(1).strip()

        # ── Total appraisal ────────────────────────────────────────
        m = re.search(r'Total\s*Appraisal\s*</td>\s*<td[^>]*>\s*\$?([\d,]+)', clean_html, re.I)
        if m:
            result['total_appraisal'] = '$' + m.group(1).strip()

        # ── Sales history ──────────────────────────────────────────
        # Find the Sales History section
        sales_section = re.search(r'Sales\s*History(.*?)(?:Parcel\s*History|</table>.*?<table)', clean_html, re.I|re.DOTALL)
        if sales_section:
            section = sales_section.group(1)
            # Each row: date, price, sale type, source code, validity code, book, page
            rows = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})\s*</td>.*?\$([\d,]+).*?</td>.*?([^<]{3,})</td>.*?(\d+)</td>.*?(\d+)</td>.*?(\d+)</td>.*?(\d+)</td>', section, re.DOTALL)
            for r in rows[:5]:
                result['sales_history'].append({
                    'date': r[0],
                    'price': '$'+r[1],
                    'type': r[2].strip(),
                    'book': r[5],
                    'page': r[6],
                })
            # Simpler fallback if above doesn't work
            if not result['sales_history']:
                dates = re.findall(r'(\d{1,2}/\d{1,2}/\d{4})', section)
                prices = re.findall(r'\$([\d,]+)', section)
                books = re.findall(r'(\d{3,})</td>\s*<td[^>]*>(\d{2,})</td>', section)
                for i in range(min(len(dates), len(prices), 5)):
                    entry = {'date': dates[i], 'price': '$'+prices[i], 'type': '', 'book': '', 'page': ''}
                    if i < len(books):
                        entry['book'] = books[i][0]
                        entry['page'] = books[i][1]
                    result['sales_history'].append(entry)

        # ── Parcel history (owner per year) ────────────────────────
        parcel_section = re.search(r'Parcel\s*History(.*?)$', clean_html, re.I|re.DOTALL)
        if parcel_section:
            section = parcel_section.group(1)
            rows = re.findall(r'(20\d\d)\s*</td>.*?(\d+)\s*</td>.*?([A-Z][A-Z\s,&.]+)\s*</td>', section, re.DOTALL)
            for r in rows[:6]:
                result['parcel_history'].append({
                    'year': r[0],
                    'tax_class': r[1],
                    'owner': r[2].strip(),
                })

        return result

    except Exception as e:
        return {'success': False, 'error': str(e), 'assessment_url': url if 'url' in dir() else ''}



# ── TITLE SEARCH VIA PLAYWRIGHT ───────────────────────────────────────────────

IDX_COUNTY_URLS = {
    "LINCOLN":    "http://129.71.206.62/Default.aspx",
    "LOGAN":      "https://loganwv.compiled-technologies.com",
    "NICHOLAS":   "http://129.71.205.250/Default.aspx",
    "GILMER":     "http://www.gilmercountywv.gov/idxsearch/",
    "HARRISON":   "http://lookup.harrisoncountywv.com/",
    "LEWIS":      "http://inquiry.lewiscountywv.org/",
    "MARSHALL":   "http://129.71.117.225/",
    "BARBOUR":    "http://129.71.117.241/WEBInquiry/Default.aspx",
    "GRANT":      "http://129.71.112.124/",
    "GREENBRIER": "http://129.71.205.208/",
    "HAMPSHIRE":  "http://129.71.205.207/idxsearch",
    "FAYETTE":    "http://129.71.202.7/",
    "DODDRIDGE":  "http://129.71.205.241/",
    "CABELL":     "http://www.recordscabellcountyclerk.org/Default.aspx",
    "JEFFERSON":  "http://documents.jeffersoncountywv.org/",
    "WIRT":       "http://records.wirtcountywv.net/",
}

# Instrument types that indicate debt/encumbrance
LIEN_TYPES = [
    "DEED OF TRUST", "TRUST DEED", "MORTGAGE", "LIEN",
    "JUDGMENT", "MECHANIC", "UCC", "TAX LIEN", "FEDERAL",
    "STATE TAX", "IRS", "ATTACHMENT"
]

# Instrument types that indicate release/satisfaction
RELEASE_TYPES = [
    "RELEASE", "SATISFACTION", "DISCHARGE", "RECONVEYANCE",
    "PARTIAL RELEASE", "FULL RELEASE"
]

# Instrument types that indicate deed/ownership transfer
DEED_TYPES = [
    "DEED", "SPECIAL WARRANTY", "GENERAL WARRANTY", "QUITCLAIM",
    "EXECUTOR", "ADMINISTRATOR", "TRUSTEE DEED", "COMMISSIONER"
]

def get_playwright_browser():
    """Launch a headless Chromium browser with minimal memory footprint."""
    from playwright.sync_api import sync_playwright
    print("[IDX] Launching Chromium browser...", flush=True)
    p = sync_playwright().start()
    browser = p.chromium.launch(
        headless=True,
        args=[
            "--no-sandbox",
            "--disable-setuid-sandbox",
            "--disable-dev-shm-usage",
            "--disable-gpu",
            "--single-process",
            "--no-zygote",
            "--disable-extensions",
            "--disable-background-networking",
            "--disable-background-timer-throttling",
            "--disable-backgrounding-occluded-windows",
            "--disable-renderer-backgrounding",
            "--disable-features=TranslateUI",
            "--disable-ipc-flooding-protection",
            "--memory-pressure-off",
            "--max_old_space_size=512",
            "--js-flags=--max-old-space-size=256",
        ]
    )
    print("[IDX] Browser launched OK", flush=True)
    return p, browser


# ── IDX TITLE SEARCH VIA PLAYWRIGHT ───────────────────────────────────────────

IDX_COUNTY_URLS = {
    "LINCOLN":    "http://129.71.206.62/Default.aspx",
    "LOGAN":      "https://loganwv.compiled-technologies.com",
    "NICHOLAS":   "http://129.71.205.250/Default.aspx",
    "GILMER":     "http://www.gilmercountywv.gov/idxsearch/Default.aspx",
    "HARRISON":   "http://lookup.harrisoncountywv.com/Default.aspx",
    "LEWIS":      "http://inquiry.lewiscountywv.org/Default.aspx",
    "MARSHALL":   "http://129.71.117.225/Default.aspx",
    "BARBOUR":    "http://129.71.117.241/WEBInquiry/Default.aspx",
    "GRANT":      "http://129.71.112.124/Default.aspx",
    "GREENBRIER": "http://129.71.205.208/Default.aspx",
    "HAMPSHIRE":  "http://129.71.205.207/idxsearch/Default.aspx",
    "FAYETTE":    "http://129.71.202.7/Default.aspx",
    "DODDRIDGE":  "http://129.71.205.241/Default.aspx",
    "CABELL":     "http://www.recordscabellcountyclerk.org/Default.aspx",
    "JEFFERSON":  "http://documents.jeffersoncountywv.org/Default.aspx",
    "WIRT":       "http://records.wirtcountywv.net/Default.aspx",
}

LIEN_TYPES = ["DEED OF TRUST","TRUST DEED","MORTGAGE","LIEN","JUDGMENT",
               "MECHANIC","UCC","TAX LIEN","FEDERAL","IRS","ATTACHMENT"]
RELEASE_TYPES = ["RELEASE","SATISFACTION","DISCHARGE","RECONVEYANCE"]
DEED_TYPES = ["DEED","WARRANTY","QUITCLAIM","EXECUTOR","ADMINISTRATOR","COMMISSIONER"]

def get_playwright_browser():
    from playwright.sync_api import sync_playwright
    print("[IDX] Launching browser...", flush=True)
    p = sync_playwright().start()
    browser = p.chromium.launch(headless=True, args=[
        "--no-sandbox","--disable-setuid-sandbox","--disable-dev-shm-usage",
        "--disable-gpu","--single-process","--no-zygote",
    ])
    print("[IDX] Browser ready", flush=True)
    return p, browser

def idx_search(page, search_type, fields, from_date="01/01/1700"):
    """
    Search IDX by mimicking keyboard navigation exactly as a human would:
    Tab 3 times to reach the search type dropdown, type to select,
    Tab to next field, type value, Tab again, type value, Enter to search.
    """
    print(f"[IDX] Keyboard search: {search_type} {fields}", flush=True)

    # Click on the page body first to ensure focus
    page.keyboard.press('Tab')
    page.wait_for_timeout(300)
    page.keyboard.press('Tab')
    page.wait_for_timeout(300)
    page.keyboard.press('Tab')
    page.wait_for_timeout(300)

    # Now type the search type - this selects from the dropdown
    # e.g. type "book" to get "Book & Page"
    type_prefix = {
        "Book & Page": "book",
        "Individual": "indiv",
        "Firm": "firm",
        "Instrument": "inst",
        "Description": "desc",
        "Date Range": "date",
        "Name": "name",
    }.get(search_type, search_type.lower()[:4])

    print(f"[IDX] Typing '{type_prefix}' to select dropdown option", flush=True)
    page.keyboard.type(type_prefix, delay=100)
    page.wait_for_timeout(1500)

    # Tab to move to the first input field (Book # or Last name)
    page.keyboard.press('Tab')
    page.wait_for_timeout(500)

    if search_type == "Book & Page":
        book = str(fields.get('book', ''))
        pg = str(fields.get('page', ''))
        print(f"[IDX] Typing book={book}", flush=True)
        page.keyboard.type(book, delay=100)
        page.wait_for_timeout(300)
        page.keyboard.press('Tab')
        page.wait_for_timeout(300)
        print(f"[IDX] Typing page={pg}", flush=True)
        page.keyboard.type(pg, delay=100)
        page.wait_for_timeout(300)

    elif search_type == "Individual":
        last = fields.get('last', '')
        first = fields.get('first', '')
        print(f"[IDX] Typing last={last}", flush=True)
        page.keyboard.type(last, delay=100)
        page.wait_for_timeout(300)
        page.keyboard.press('Tab')
        page.wait_for_timeout(300)
        print(f"[IDX] Typing first={first}", flush=True)
        page.keyboard.type(first, delay=100)
        page.wait_for_timeout(300)

    # Press Enter to search
    print("[IDX] Pressing Enter to search", flush=True)
    page.keyboard.press('Enter')
    page.wait_for_load_state('domcontentloaded', timeout=20000)
    page.wait_for_timeout(4000)

    # Take screenshot and save HTML for debugging
    try:
        page.screenshot(path="/tmp/idx_screenshot.png", full_page=False)
        print("[IDX] Screenshot saved to /tmp/idx_screenshot.png", flush=True)
    except Exception as e:
        print(f"[IDX] Screenshot failed: {e}", flush=True)

    try:
        text = page.inner_text('body')
        print(f"[IDX] Results page preview: {text[:800]}", flush=True)
    except: pass

    html = page.content()
    # Count dxgvDataRow before parsing
    import re
    dr = re.findall(r'dxgvDataRow', html)
    print(f"[IDX] dxgvDataRow count in HTML: {len(dr)}", flush=True)

    return parse_idx_results(html)

def parse_idx_results(html):
    """Parse IDX DevExpress grid results."""
    import re, html as html_mod
    records = []

    # The DevExpress grid rows have class dxgvDataRow
    # Extract them specifically
    data_rows = re.findall(
        r'<tr[^>]*class="[^"]*dxgvDataRow[^"]*"[^>]*>(.*?)</tr>',
        html, re.DOTALL|re.IGNORECASE
    )
    print(f"[IDX] Found {len(data_rows)} dxgvDataRow rows", flush=True)

    # Also find header row
    header_rows = re.findall(
        r'<tr[^>]*class="[^"]*dxgvHeader[^"]*"[^>]*>(.*?)</tr>',
        html, re.DOTALL|re.IGNORECASE
    )
    header = []
    for hr in header_rows:
        cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', hr, re.DOTALL|re.IGNORECASE)
        cells = [re.sub(r'<[^>]+>','',c).strip() for c in cells]
        cells = [html_mod.unescape(' '.join(c.split())) for c in cells]
        cells = [c for c in cells if c]
        if cells:
            header = cells
            print(f"[IDX] Grid header: {cells}", flush=True)
            break

    for row_html in data_rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row_html, re.DOTALL|re.IGNORECASE)
        cells = [re.sub(r'<[^>]+>','',c).strip() for c in cells]
        cells = [html_mod.unescape(' '.join(c.split())) for c in cells]
        cells = [c for c in cells if c]
        if not cells: continue

        if header and len(header) == len(cells):
            rec = {header[i].lower().replace(' ','_'): cells[i] for i in range(len(cells))}
        else:
            rec = {'raw': cells}
        records.append(rec)
        print(f"[IDX] Result: {cells}", flush=True)

    # If no dxgvDataRow found, fall back to any table rows with deed data
    if not records:
        print("[IDX] No dxgvDataRow found, trying fallback parse", flush=True)
        all_rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL|re.IGNORECASE)
        SKIP = {'SUN','MON','TUE','WED','THU','FRI','SAT'}
        for row in all_rows:
            cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL|re.IGNORECASE)
            cells = [re.sub(r'<[^>]+>','',c).strip() for c in cells]
            cells = [html_mod.unescape(' '.join(c.split())) for c in cells]
            cells = [c for c in cells if c]
            if not cells or len(cells) < 3: continue
            if set(c.upper() for c in cells) & SKIP: continue
            if all(c.isdigit() and int(c)<=31 for c in cells if c): continue
            # Must have a date pattern to be a deed record
            if not any(re.search(r'\d{1,2}/\d{1,2}/\d{4}', c) for c in cells): continue
            rec = {'raw': cells}
            records.append(rec)
            print(f"[IDX] Fallback row: {cells}", flush=True)

    print(f"[IDX] Total records: {len(records)}", flush=True)
    return records

def do_title_search(county_name, deed_book, deed_page, current_owner_name, years_back=25):
    """Full title search using Playwright."""
    county = county_name.upper().replace(" COUNTY","").strip()
    url = IDX_COUNTY_URLS.get(county)
    if not url:
        return {"success": False, "error": f"No IDX configured for {county}"}
    if not deed_book or not deed_page:
        return {"success": False, "error": "Deed book and page required"}

    print(f"[IDX] Title search: {county} Book {deed_book} Page {deed_page}", flush=True)
    cutoff_year = datetime.now().year - years_back

    report = {
        "success": True, "county": county,
        "starting_book": deed_book, "starting_page": deed_page,
        "current_owner": current_owner_name,
        "chain_of_title": [], "open_liens": [],
        "released_liens": [], "all_instruments": [], "errors": []
    }

    try:
        p, browser = get_playwright_browser()
        ctx = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
        page = ctx.new_page()
        page.set_default_timeout(20000)

        # Load page
        print(f"[IDX] Loading {url}", flush=True)
        page.goto(url, wait_until='domcontentloaded', timeout=25000)
        page.wait_for_timeout(8000)  # Wait for JS to fully initialize

        # Search by Book & Page
        book_records = idx_search(page, "Book & Page", {
            'book': deed_book, 'page': deed_page
        })
        print(f"[IDX] Book/Page results: {len(book_records)}", flush=True)
        report['chain_of_title'].append({
            "owner": current_owner_name,
            "deed_book": deed_book, "deed_page": deed_page,
            "instruments": book_records
        })
        report['all_instruments'].extend(book_records)

        # Reload page for name search
        print("[IDX] Reloading for name search...", flush=True)
        page.goto(url, wait_until='domcontentloaded', timeout=25000)
        page.wait_for_timeout(5000)

        # Search by name
        # WV names are stored as "FIRSTNAME LASTNAME" or "LASTNAME FIRSTNAME"
        # Assessment shows "KEENEY DON" meaning KEENEY=last, DON=first
        parts = current_owner_name.strip().split()
        last = parts[0]   # First word is last name in WV records
        first = parts[1] if len(parts) > 1 else ""
        name_records = idx_search(page, "Individual", {
            'last': last, 'first': first
        }, from_date=f"01/01/{cutoff_year}")
        print(f"[IDX] Name results: {len(name_records)}", flush=True)

        for rec in name_records:
            rec_text = ' '.join(str(v) for v in rec.values()).upper()
            is_lien = any(lt in rec_text for lt in LIEN_TYPES)
            is_release = any(rt in rec_text for rt in RELEASE_TYPES)
            if is_lien or is_release:
                report['all_instruments'].append(rec)
                if is_release:
                    report['released_liens'].append(rec)
                elif is_lien:
                    report['open_liens'].append({"status":"POSSIBLY OPEN","details":rec})

        browser.close()
        p.stop()
        return report

    except Exception as e:
        import traceback
        print(f"[IDX] Error: {e}", flush=True)
        traceback.print_exc()
        return {"success": False, "error": str(e)}



def sync_wvsao_dates():
    """
    Fetch all auction dates from WVSAO.
    Strategy: 
    1. Fetch main page - get page 1 dates + total pages + viewstate
    2. For pages 2-N: POST with Page$N argument (ASP.NET ListView pager format)
    3. Fallback: parse handouts list for county names, dates from listings
    """
    import urllib.request as ur
    import urllib.parse
    import re

    BASE = "https://www.wvsao.gov/CountyCollections/Default"
    HEADERS = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
        'Referer': BASE,
    }

    def extract_hidden(html):
        fields = {}
        for m in re.finditer(r'<input[^>]+type="hidden"[^>]*>', html, re.I):
            nm = re.search(r'name="([^"]+)"', m.group(0))
            vl = re.search(r'value="([^"]*)"', m.group(0))
            if nm:
                fields[nm.group(1)] = vl.group(1) if vl else ''
        return fields

    def parse_dates_from_html(html):
        """Extract date+county pairs from raw HTML."""
        # Strip tags
        clean = re.sub(r'<[^>]+>', ' ', html)
        clean = re.sub(r'\s+', ' ', clean)
        
        results = {}
        # Match: Date: MM/DD/YYYY ... County: XXXXX COUNTY
        # The page shows Date, Time, County, Seller, Location in order
        rx = re.compile(
            r'Date:\s*(\d{1,2}/\d{1,2}/\d{4})\s+Time:[^C]*County:\s*([A-Z][A-Z\s]{2,}?COUNTY)',
            re.I
        )
        for m in rx.finditer(clean):
            p = m.group(1).strip().split('/')
            if len(p) == 3:
                iso = f"{p[2]}-{p[0].zfill(2)}-{p[1].zfill(2)}"
                county_raw = m.group(2).strip()
                county = re.sub(r'\s*COUNTY\s*$', '', county_raw, flags=re.I).strip().title()
                if county not in results:
                    results[county] = iso
                    print(f"[WVSAO] Found: {county} = {iso}", flush=True)
        return results

    def do_get(url):
        req = ur.Request(url, headers=HEADERS)
        with ur.urlopen(req, timeout=15) as r:
            return r.read().decode('utf-8', errors='replace')

    def do_post(url, fields):
        data = urllib.parse.urlencode(fields).encode()
        h = {**HEADERS, 'Content-Type': 'application/x-www-form-urlencoded'}
        req = ur.Request(url, data=data, headers=h)
        with ur.urlopen(req, timeout=15) as r:
            return r.read().decode('utf-8', errors='replace')

    try:
        found = {}

        # Page 1
        html = do_get(BASE)
        found.update(parse_dates_from_html(html))
        print(f"[WVSAO] Page 1: {len(found)} counties", flush=True)

        # Get total pages from "Page 1 of N (X results)"
        clean = re.sub(r'<[^>]+>', ' ', html)
        pm = re.search(r'Page\s+1\s+of\s+(\d+)', clean, re.I)
        total_pages = int(pm.group(1)) if pm else 1
        rm = re.search(r'\((\d+)\s+results\)', clean, re.I)
        total_results = int(rm.group(1)) if rm else 0
        print(f"[WVSAO] {total_results} auctions across {total_pages} pages", flush=True)

        # Get ViewState
        vs = extract_hidden(html)

        # Pages 2 to N
        for pg in range(2, total_pages + 1):
            pg_html = None
            # ASP.NET ListView uses "Page$N" for page N
            for target, arg in [
                ('ctl00$FixedWidthContent$ListView1', f'Page${pg}'),
                ('ctl00$FixedWidthContent$ListView1', f'MoveToPage;{pg-1}'),
            ]:
                try:
                    fields = dict(vs)
                    fields['__EVENTTARGET'] = target
                    fields['__EVENTARGUMENT'] = arg
                    pg_html = do_post(BASE, fields)
                    new_dates = parse_dates_from_html(pg_html)
                    print(f"[WVSAO] Page {pg} (arg={arg}): {new_dates}", flush=True)
                    if new_dates:
                        found.update(new_dates)
                        vs = extract_hidden(pg_html)
                        break
                    elif len(pg_html) > 50000:
                        # Page loaded but no new dates (already seen or different format)
                        vs = extract_hidden(pg_html)
                        break
                except Exception as e:
                    print(f"[WVSAO] Page {pg} {arg} error: {e}", flush=True)

        print(f"[WVSAO] Complete: {len(found)} counties - {found}", flush=True)
        return {
            "success": True,
            "dates": found,
            "count": len(found),
            "total_pages": total_pages,
            "total_auctions": total_results
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"success": False, "error": str(e)}



async def scrape_og_intel(owner_name, county, district, map_num, parcel, min_bid, description):
    """
    Scrape WV Assessment portal + WVDEP well database using Playwright.
    Returns dict with mineral assessment data, well data, and AI analysis.
    """
    import asyncio
    from playwright.async_api import async_playwright

    results = {
        "owner": owner_name,
        "county": county,
        "assessments": [],      # All property records for this owner
        "mineral_parcels": [],  # Specifically mineral/O&G parcels
        "wells": [],            # Active wells in same county/district
        "raw_errors": []
    }

    # County number mapping for mapwv.gov assessment portal
    COUNTY_NUMS = {
        "BARBOUR":"1","BERKELEY":"2","BOONE":"3","BRAXTON":"4","BROOKE":"5",
        "CABELL":"6","CALHOUN":"7","CLAY":"8","DODDRIDGE":"9","FAYETTE":"10",
        "GILMER":"11","GRANT":"12","GREENBRIER":"13","HAMPSHIRE":"14","HANCOCK":"15",
        "HARDY":"16","HARRISON":"17","JACKSON":"18","JEFFERSON":"19","KANAWHA":"20",
        "LEWIS":"21","LINCOLN":"22","LOGAN":"23","MARION":"24","MARSHALL":"25",
        "MASON":"26","MCDOWELL":"27","MERCER":"28","MINERAL":"29","MINGO":"30",
        "MONONGALIA":"31","MONROE":"32","MORGAN":"33","NICHOLAS":"34","OHIO":"35",
        "PENDLETON":"36","PLEASANTS":"37","POCAHONTAS":"38","PRESTON":"39","PUTNAM":"40",
        "RALEIGH":"41","RANDOLPH":"42","RITCHIE":"43","ROANE":"44","SUMMERS":"45",
        "TAYLOR":"46","TUCKER":"47","TYLER":"48","UPSHUR":"49","WAYNE":"50",
        "WEBSTER":"51","WETZEL":"52","WIRT":"53","WOOD":"54","WYOMING":"55"
    }

    county_upper = county.upper().replace(" COUNTY","").strip()
    county_num = COUNTY_NUMS.get(county_upper, "")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
        ctx = await browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")

        # ── STEP 1: WV Assessment Portal — search by owner name ─────────────────
        try:
            page = await ctx.new_page()
            print(f"[OG-INTEL] Loading assessment portal for {owner_name} in {county}", flush=True)
            await page.goto("https://www.mapwv.gov/assessment/Assessment", timeout=30000)
            await page.wait_for_load_state("networkidle", timeout=15000)

            # Set county if we have a number
            if county_num:
                await page.select_option("select[name*='county'], select[id*='county'], #County, #ddlCounty", 
                    value=county_num, timeout=5000)

            # Fill owner name - try common input field names
            for selector in ["#OwnerName", "input[name*='owner']", "input[placeholder*='owner' i]", 
                             "input[name*='Owner']", "#txtOwnerName"]:
                try:
                    await page.fill(selector, owner_name, timeout=3000)
                    print(f"[OG-INTEL] Filled owner name in {selector}", flush=True)
                    break
                except:
                    continue

            # Click search
            for sel in ["input[type=submit]", "button[type=submit]", "#btnSearch", 
                        "input[value*='Search' i]", "button:has-text('Search')"]:
                try:
                    await page.click(sel, timeout=3000)
                    print(f"[OG-INTEL] Clicked search via {sel}", flush=True)
                    break
                except:
                    continue

            await page.wait_for_load_state("networkidle", timeout=15000)
            await page.wait_for_timeout(2000)

            # Parse results table
            html = await page.content()
            rows = await page.query_selector_all("table tr, .result-row, tr[class*='row']")
            print(f"[OG-INTEL] Found {len(rows)} rows in assessment results", flush=True)

            for row in rows[:50]:  # limit to 50
                try:
                    cells = await row.query_selector_all("td")
                    if len(cells) < 3:
                        continue
                    texts = []
                    for cell in cells:
                        t = (await cell.inner_text()).strip()
                        texts.append(t)

                    row_text = " | ".join(texts)
                    print(f"[OG-INTEL] Row: {row_text[:150]}", flush=True)

                    # Detect mineral/O&G parcels
                    is_mineral = any(kw in row_text.upper() for kw in [
                        "MINERAL","OIL","GAS","O&G","ROYALT","MIN ","NATURAL GAS",
                        "PRODUCING","MARCELLUS","UTICA","COAL","SUBSURFACE"
                    ])

                    record = {"cells": texts, "raw": row_text, "is_mineral": is_mineral}
                    results["assessments"].append(record)
                    if is_mineral:
                        results["mineral_parcels"].append(record)
                except Exception as e:
                    continue

        except Exception as e:
            msg = f"Assessment portal error: {str(e)}"
            print(f"[OG-INTEL] {msg}", flush=True)
            results["raw_errors"].append(msg)

        # ── STEP 2: WVDEP Well Database ─────────────────────────────────────────
        try:
            page2 = await ctx.new_page()
            print(f"[OG-INTEL] Loading WVDEP well DB for {county}", flush=True)
            await page2.goto("https://tagis.dep.wv.gov/oog/", timeout=30000)
            await page2.wait_for_load_state("networkidle", timeout=15000)

            # Select county
            try:
                await page2.select_option("select[name*='county' i], #county, #ddlCounty",
                    label=county_upper.title(), timeout=5000)
            except:
                pass

            # Select Active wells + Gas Production
            try:
                await page2.select_option("select[name*='status' i], #wellstatus",
                    label="Active Well", timeout=3000)
            except:
                pass
            try:
                await page2.select_option("select[name*='use' i], #welluse",
                    label="Gas Production", timeout=3000)
            except:
                pass

            # Select Horizontal 6A (Marcellus/Utica)
            try:
                await page2.select_option("select[name*='type' i], #permittype",
                    label="Horizontal 6A Well", timeout=3000)
            except:
                pass

            # Search
            for sel in ["input[type=submit]", "input[value*='Search' i]", "#btnSearch"]:
                try:
                    await page2.click(sel, timeout=3000)
                    break
                except:
                    continue

            await page2.wait_for_load_state("networkidle", timeout=20000)
            await page2.wait_for_timeout(2000)

            rows2 = await page2.query_selector_all("table tr")
            print(f"[OG-INTEL] Found {len(rows2)} well rows for {county}", flush=True)

            for row in rows2[:30]:
                try:
                    cells = await row.query_selector_all("td")
                    if len(cells) < 3:
                        continue
                    texts = [(await c.inner_text()).strip() for c in cells]
                    row_text = " | ".join(texts)
                    if any(kw in row_text.upper() for kw in ["GAS","OIL","MARCELLUS","HORIZONTAL","ACTIVE"]):
                        results["wells"].append({"cells": texts, "raw": row_text})
                        print(f"[OG-INTEL] Well: {row_text[:120]}", flush=True)
                except:
                    continue

        except Exception as e:
            msg = f"WVDEP well error: {str(e)}"
            print(f"[OG-INTEL] {msg}", flush=True)
            results["raw_errors"].append(msg)

        await browser.close()

    return results


def run_og_intel(owner_name, county, district, map_num, parcel, min_bid, description):
    """Synchronous wrapper for the async scraper."""
    import asyncio
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        data = loop.run_until_complete(
            scrape_og_intel(owner_name, county, district, map_num, parcel, min_bid, description)
        )
        loop.close()
        return data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e), "owner": owner_name}


def build_og_assessment(scraped, owner_name, county, district, min_bid, description):
    """
    Feed scraped data to Claude for plain-English O&G intelligence assessment.
    """
    import anthropic
    client = anthropic.Anthropic()

    # Formation tier knowledge
    FORMATION_TIERS = {
        "MARSHALL": ("Tier 1", "Marcellus + Utica sweet spot — highest production in WV"),
        "WETZEL": ("Tier 1", "Marcellus sweet spot — top producing county"),
        "TYLER": ("Tier 1", "Marcellus + Utica high production, many active H6A wells"),
        "DODDRIDGE": ("Tier 1", "Strong Marcellus production, active drilling"),
        "RITCHIE": ("Tier 1", "Marcellus producer, active horizontal drilling"),
        "PLEASANTS": ("Tier 2", "Marcellus present, moderate production"),
        "WOOD": ("Tier 2", "Marcellus present, moderate production"),
        "KANAWHA": ("Tier 2", "Marcellus present, active drilling"),
        "LINCOLN": ("Tier 2", "Marcellus present, some active wells"),
        "ROANE": ("Tier 2", "Marcellus + conventional O&G history"),
        "CALHOUN": ("Tier 2", "Conventional O&G + Marcellus fringe"),
        "WIRT": ("Tier 2", "Conventional O&G history"),
        "JACKSON": ("Tier 2", "Marcellus fringe, conventional O&G"),
        "PUTNAM": ("Tier 2", "Marcellus present, some drilling"),
        "BRAXTON": ("Tier 2", "Conventional O&G + Marcellus fringe"),
        "NICHOLAS": ("Tier 2", "Conventional O&G history"),
        "LOGAN": ("Tier 2", "Conventional O&G + coal"),
    }
    county_upper = county.upper().replace(" COUNTY","").strip()
    tier, tier_desc = FORMATION_TIERS.get(county_upper, ("Tier 3", "Limited Marcellus/Utica production expected"))

    mineral_found = len(scraped.get("mineral_parcels", []))
    total_assessed = len(scraped.get("assessments", []))
    wells_found = len(scraped.get("wells", []))

    assessment_summary = chr(10).join([
        r["raw"][:200] for r in scraped.get("assessments", [])[:10]
    ]) or "No assessment data retrieved"

    well_summary = chr(10).join([
        r["raw"][:200] for r in scraped.get("wells", [])[:10]
    ]) or "No active well data retrieved"

    prompt = f"""You are an expert West Virginia oil and gas mineral rights analyst. Analyze this tax lien property and provide an intelligence assessment.

PROPERTY DATA:
- Owner: {owner_name}
- County: {county}
- District: {district}
- Description: {description}
- Minimum Bid: {min_bid}

FORMATION INTELLIGENCE:
- {county_upper} County Formation Tier: {tier}
- Assessment: {tier_desc}

COUNTY ASSESSOR DATA (pulled live from mapwv.gov):
{assessment_summary}

ACTIVE WELLS IN COUNTY (from WVDEP):
{well_summary}

IMPORTANT CONTEXT:
- In WV, if minerals are PRODUCING, the operator reports royalties to the State Tax Division
- The assessor's assessed value for producing minerals = 1.5x to 7x the annual royalty income
- A 2-year delay exists between production start and assessment update
- Horizontal 6A (H6A) wells = Marcellus/Utica shale horizontal wells = highest royalty producers
- "MIN" in the description = mineral interest (not surface rights)
- Fractions like "1/8 OF 154 AC" = royalty fraction of acreage

Provide a structured assessment with:
1. ROYALTY STATUS: Are these minerals likely producing? Evidence from assessor data?
2. FORMATION RISK: Based on county tier and well data
3. ESTIMATED VALUE: If producing, what annual royalty range is plausible?
4. OPERATOR INTEL: Any O&G companies identifiable from the well data?
5. RECOMMENDATION: Priority (HIGH/MEDIUM/LOW) and why
6. RISK FLAGS: Any issues (old wells, plugged wells, no production evidence)

Be specific and data-driven. Reference actual numbers from the scraped data where available."""

    msg = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=1000,
        messages=[{"role": "user", "content": prompt}]
    )
    return msg.content[0].text



# ── SHERIFF TAX LOOKUP ────────────────────────────────────────────────────────
# Software Systems Inc system used by most WV counties
# URL: http://{county}.softwaresystems.com/
# Search by ticket number → returns appraised value, assessed value, tax amount

SHERIFF_TAX_URLS = {
    "LINCOLN": "http://lincoln.softwaresystems.com",
    "PUTNAM": "http://putnam.softwaresystems.com",
    "KANAWHA": "http://kanawha.softwaresystems.com",
    "CLAY": "http://clay.softwaresystems.com",
    "NICHOLAS": "http://nicholas.softwaresystems.com",
    "BRAXTON": "http://braxton.softwaresystems.com",
    "WEBSTER": "http://webster.softwaresystems.com",
    "GILMER": "http://gilmer.softwaresystems.com",
    "CALHOUN": "http://calhoun.softwaresystems.com",
    "ROANE": "http://roane.softwaresystems.com",
    "LOGAN": "http://logan.softwaresystems.com",
    "MARSHALL": "http://marshall.softwaresystems.com",
    "WETZEL": "http://wetzel.softwaresystems.com",
    "TYLER": "http://tyler.softwaresystems.com",
    "DODDRIDGE": "http://doddridge.softwaresystems.com",
    "WIRT": "http://wirt.softwaresystems.com",
    "JACKSON": "http://jackson.softwaresystems.com",
    "WOOD": "http://wood.softwaresystems.com",
}

async def scrape_sheriff_tax(county, ticket, district_num=None, map_num=None, parcel=None, tax_year=None):
    """
    Scrape sheriff tax office for appraised value, assessed value, and actual tax.
    Returns dict with financial data for ROI calculation.
    """
    from playwright.async_api import async_playwright
    import re

    county_upper = county.upper().replace(" COUNTY","").strip()
    base_url = SHERIFF_TAX_URLS.get(county_upper)
    if not base_url:
        return {"error": f"No sheriff URL for {county_upper}", "supported": list(SHERIFF_TAX_URLS.keys())}

    result = {
        "county": county_upper,
        "ticket": ticket,
        "appraised_value": None,
        "assessed_value": None, 
        "tax_amount": None,
        "tax_year": None,
        "owner": None,
        "district": None,
        "map": None,
        "parcel": None,
        "description": None,
        "status": None,
        "raw_rows": [],
        "error": None
    }

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
        ctx = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        )
        page = await ctx.new_page()

        try:
            print(f"[SHERIFF] Loading {base_url}", flush=True)
            await page.goto(base_url + "/index.html", timeout=30000, wait_until="domcontentloaded")
            await page.wait_for_timeout(1000)

            # Fill ticket number
            ticket_filled = False
            for sel in ["input[name=TICKET]", "input[name=TPTICK]", "input[name*=ticket i]"]:
                try:
                    await page.fill(sel, str(ticket), timeout=3000)
                    ticket_filled = True
                    print(f"[SHERIFF] Ticket filled via {sel}", flush=True)
                    break
                except:
                    pass

            # Fill tax year if provided
            if tax_year:
                for sel in ["input[name=TAXYR]", "input[name=TPTYR]", "input[name*=year i]"]:
                    try:
                        await page.fill(sel, str(tax_year), timeout=2000)
                        break
                    except:
                        pass

            # Set real estate type
            try:
                await page.select_option("select[name=TXTYPE]", value="R", timeout=2000)
            except:
                pass

            # Submit search
            for sel in ["input[type=submit]", "input[value=Search]", "input[value=Search i]",
                        "button[type=submit]", "input[name=SEARCH]"]:
                try:
                    await page.click(sel, timeout=3000)
                    print(f"[SHERIFF] Search submitted via {sel}", flush=True)
                    break
                except:
                    pass

            await page.wait_for_load_state("domcontentloaded", timeout=15000)
            await page.wait_for_timeout(2000)

            # Get results page content
            body = await page.inner_text("body")
            print(f"[SHERIFF] Results page text (first 500):", flush=True)
            print(body[:500], flush=True)

            # Parse tables
            tables = await page.query_selector_all("table")
            all_rows = []
            for table in tables:
                rows = await table.query_selector_all("tr")
                for row in rows:
                    cells = await row.query_selector_all("td, th")
                    if cells:
                        texts = [(await c.inner_text()).strip() for c in cells]
                        if any(t for t in texts):
                            all_rows.append(texts)
                            print(f"[SHERIFF] Row: {texts}", flush=True)

            result["raw_rows"] = all_rows

            # Look for a link to the specific ticket and click it
            links = await page.query_selector_all("a")
            for link in links:
                href = await link.get_attribute("href") or ""
                txt = (await link.inner_text()).strip()
                if str(ticket) in href or str(ticket) in txt:
                    print(f"[SHERIFF] Clicking ticket link: {href}", flush=True)
                    await link.click()
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await page.wait_for_timeout(2000)
                    body2 = await page.inner_text("body")
                    print(f"[SHERIFF] Ticket detail (first 800):", flush=True)
                    print(body2[:800], flush=True)

                    # Parse the detail page
                    tables2 = await page.query_selector_all("table")
                    for table in tables2:
                        rows2 = await table.query_selector_all("tr")
                        for row in rows2:
                            cells = await row.query_selector_all("td, th")
                            texts = [(await c.inner_text()).strip() for c in cells]
                            if any(t for t in texts):
                                result["raw_rows"].append(texts)
                                row_text = " | ".join(texts).upper()

                                # Extract financial values
                                if "APPRAISED" in row_text or "APPRAIS" in row_text:
                                    for t in texts:
                                        m = re.search(r"\\$?([\d,]+\.?\d*)", t.replace(",",""))
                                        if m and float(m.group(1)) > 0:
                                            result["appraised_value"] = float(m.group(1))
                                if "ASSESSED" in row_text:
                                    for t in texts:
                                        m = re.search(r"\\$?([\d,]+\.?\d*)", t.replace(",",""))
                                        if m and float(m.group(1)) > 0:
                                            result["assessed_value"] = float(m.group(1))
                                if "TAX" in row_text and ("AMOUNT" in row_text or "DUE" in row_text or "TOTAL" in row_text):
                                    for t in texts:
                                        m = re.search(r"\\$?([\d,]+\.?\d*)", t.replace(",",""))
                                        if m and float(m.group(1)) > 0:
                                            result["tax_amount"] = float(m.group(1))
                                if "OWNER" in row_text or "NAME" in row_text:
                                    for i, t in enumerate(texts):
                                        if "OWNER" in t.upper() or "NAME" in t.upper():
                                            if i+1 < len(texts) and texts[i+1].strip():
                                                result["owner"] = texts[i+1].strip()
                    break

            # If we didn't navigate to detail, try to parse results page directly
            if not result["appraised_value"]:
                for row in all_rows:
                    row_text = " | ".join(row).upper()
                    if "APPRAISED" in row_text:
                        for t in row:
                            m = re.search(r"[\d,]+\.?\d*", t.replace(",",""))
                            if m:
                                try: result["appraised_value"] = float(m.group())
                                except: pass

        except Exception as e:
            import traceback
            traceback.print_exc()
            result["error"] = str(e)

        await browser.close()

    # Calculate ROI if we have financial data
    if result["appraised_value"]:
        av = result["appraised_value"]
        # WV mineral royalty formula: appraised = 1.5x to 7x annual royalty
        result["est_annual_royalty_low"] = round(av / 7, 2)
        result["est_annual_royalty_high"] = round(av / 1.5, 2)
        result["est_annual_royalty_mid"] = round((av/7 + av/1.5) / 2, 2)

    return result


def run_sheriff_lookup(county, ticket, district_num=None, map_num=None, parcel=None, tax_year=None):
    """Synchronous wrapper."""
    import asyncio
    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        data = loop.run_until_complete(
            scrape_sheriff_tax(county, ticket, district_num, map_num, parcel, tax_year)
        )
        loop.close()
        return data
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"error": str(e), "county": county, "ticket": ticket}

# ─────────────────────────────────────────────────────────────────────────────

"""
Bulletproof O&G Assessment Engine
Runs on Render, called via /og-assess endpoint
Multiple data sources with automatic fallbacks
"""

import asyncio, re, json
from playwright.async_api import async_playwright

# ── FORMATION INTELLIGENCE (always available) ─────────────────────────────────
FORMATION_DATA = {
    "MARSHALL": {
        "tier": 1, "marcellus": "PRIME", "utica": "PRIME",
        "desc": "Top Marcellus+Utica producer in WV. Highest royalty checks in state.",
        "active_operators": ["EQT", "CNX Resources", "Southwestern Energy", "Equinor"],
        "avg_royalty_per_acre": 850,  # $/acre/year estimate for active Marcellus
        "drilling_outlook": "VERY ACTIVE - multiple H6A permits 2023-2025"
    },
    "WETZEL": {
        "tier": 1, "marcellus": "PRIME", "utica": "STRONG",
        "desc": "Top 2 Marcellus county. Very active horizontal drilling.",
        "active_operators": ["EQT", "Southwestern Energy", "Antero Resources"],
        "avg_royalty_per_acre": 720,
        "drilling_outlook": "VERY ACTIVE"
    },
    "TYLER": {
        "tier": 1, "marcellus": "PRIME", "utica": "PRIME",
        "desc": "Top Marcellus producer. Leading Utica county per WVGES 2022.",
        "active_operators": ["EQT", "Southwestern Energy", "Tug Hill Operating"],
        "avg_royalty_per_acre": 680,
        "drilling_outlook": "ACTIVE - continued H6A development"
    },
    "DODDRIDGE": {
        "tier": 1, "marcellus": "PRIME", "utica": "STRONG",
        "desc": "Strong Marcellus formation. Active horizontal drilling.",
        "active_operators": ["EQT", "Antero Resources", "CNX"],
        "avg_royalty_per_acre": 590,
        "drilling_outlook": "ACTIVE"
    },
    "RITCHIE": {
        "tier": 1, "marcellus": "STRONG", "utica": "MODERATE",
        "desc": "Productive Marcellus area. Long conventional O&G history.",
        "active_operators": ["EQT", "Diversified Energy"],
        "avg_royalty_per_acre": 420,
        "drilling_outlook": "MODERATE - conventional plus some Marcellus"
    },
    "PLEASANTS": {
        "tier": 2, "marcellus": "STRONG", "utica": "MODERATE",
        "desc": "Marcellus present. Active conventional and unconventional production.",
        "active_operators": ["EQT", "Diversified Energy", "Cabot/Coterra"],
        "avg_royalty_per_acre": 380,
        "drilling_outlook": "MODERATE"
    },
    "WOOD": {
        "tier": 2, "marcellus": "MODERATE", "utica": "MODERATE",
        "desc": "Conventional O&G history. Some Marcellus activity.",
        "active_operators": ["Diversified Energy", "Cabot/Coterra"],
        "avg_royalty_per_acre": 280,
        "drilling_outlook": "MODERATE - mostly conventional"
    },
    "WIRT": {
        "tier": 2, "marcellus": "MODERATE", "utica": "LOW",
        "desc": "Long conventional O&G history. Some newer Marcellus permits.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 220,
        "drilling_outlook": "LOW-MODERATE"
    },
    "JACKSON": {
        "tier": 2, "marcellus": "MODERATE", "utica": "LOW",
        "desc": "Conventional O&G area. Limited Marcellus development.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 200,
        "drilling_outlook": "LOW-MODERATE"
    },
    "KANAWHA": {
        "tier": 2, "marcellus": "MODERATE", "utica": "LOW",
        "desc": "Active area with mix of conventional and Marcellus.",
        "active_operators": ["EQT", "Diversified Energy"],
        "avg_royalty_per_acre": 260,
        "drilling_outlook": "MODERATE"
    },
    "PUTNAM": {
        "tier": 2, "marcellus": "MODERATE", "utica": "LOW",
        "desc": "Some Marcellus activity. Conventional O&G history.",
        "active_operators": ["EQT", "Diversified Energy"],
        "avg_royalty_per_acre": 230,
        "drilling_outlook": "LOW-MODERATE"
    },
    "LINCOLN": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Marcellus fringe area. Active conventional production especially Guyan Gas field.",
        "active_operators": ["Guyan International", "Argus Energy", "Diversified Energy"],
        "avg_royalty_per_acre": 180,
        "drilling_outlook": "LOW - mostly conventional, Guyan Gas field active",
        "special": "GUYAN GAS FIELD active in Sheridan/Jefferson districts - high conventional O&G"
    },
    "ROANE": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Conventional O&G. Limited Marcellus.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 160,
        "drilling_outlook": "LOW"
    },
    "CALHOUN": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Conventional O&G. Some older production.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 140,
        "drilling_outlook": "LOW"
    },
    "BRAXTON": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Conventional O&G history. Some newer drilling.",
        "active_operators": ["Diversified Energy", "EQT"],
        "avg_royalty_per_acre": 150,
        "drilling_outlook": "LOW-MODERATE"
    },
    "NICHOLAS": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Conventional O&G. Limited Marcellus presence.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 130,
        "drilling_outlook": "LOW"
    },
    "CLAY": {
        "tier": 3, "marcellus": "MINIMAL", "utica": "NONE",
        "desc": "Limited O&G activity. Not a primary formation area.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 80,
        "drilling_outlook": "VERY LOW"
    },
    "LOGAN": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Coal and conventional O&G area. Some gas production.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 140,
        "drilling_outlook": "LOW"
    },
    "MINGO": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Coal and conventional O&G.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 120,
        "drilling_outlook": "LOW"
    },
    "WAYNE": {
        "tier": 2, "marcellus": "FRINGE", "utica": "LOW",
        "desc": "Some conventional O&G. Limited Marcellus.",
        "active_operators": ["Diversified Energy"],
        "avg_royalty_per_acre": 110,
        "drilling_outlook": "LOW"
    },
    "GILMER": {
        "tier": 3, "marcellus": "MINIMAL", "utica": "NONE",
        "desc": "Limited O&G. Some conventional production.",
        "active_operators": [],
        "avg_royalty_per_acre": 90,
        "drilling_outlook": "VERY LOW"
    },
    "WEBSTER": {
        "tier": 3, "marcellus": "MINIMAL", "utica": "NONE",
        "desc": "Limited O&G. Remote mountainous terrain.",
        "active_operators": [],
        "avg_royalty_per_acre": 70,
        "drilling_outlook": "VERY LOW"
    },
}

# WV county levy rates (approximate, per $100 assessed value)
# Class 3 = non-owner occupied outside municipality (minerals fall here)
COUNTY_LEVY_RATES = {
    "LINCOLN": 0.7234, "PUTNAM": 0.6890, "KANAWHA": 0.7012,
    "CLAY": 0.7145, "NICHOLAS": 0.6923, "BRAXTON": 0.7056,
    "WEBSTER": 0.7234, "GILMER": 0.6789, "CALHOUN": 0.7123,
    "ROANE": 0.7045, "LOGAN": 0.7156, "MARSHALL": 0.6834,
    "WETZEL": 0.7012, "TYLER": 0.6978, "DODDRIDGE": 0.6845,
    "WIRT": 0.6923, "JACKSON": 0.7034, "WOOD": 0.6912,
    "RITCHIE": 0.6945, "PLEASANTS": 0.6867, "WAYNE": 0.7123,
    "MINGO": 0.7234, "DEFAULT": 0.70
}

# Description signal analysis
def analyze_description(desc, name):
    """Extract key signals from legal description and owner name."""
    desc_up = (desc or "").upper()
    name_up = (name or "").upper()
    signals = []
    priority = "LOW"
    
    # Highest value signals
    if "ROYALTY INT" in desc_up or "ROYALTY INT" in name_up:
        signals.append({"type": "ROYALTY_INTEREST", "weight": 10,
            "note": "Property described as ROYALTY INTEREST - currently receiving checks"})
        priority = "HIGH"
    
    if "GUYAN GAS" in desc_up or "GUYAN GAS" in name_up:
        signals.append({"type": "GUYAN_GAS_FIELD", "weight": 9,
            "note": "Guyan Gas Field - active conventional gas producer in Lincoln County"})
        priority = "HIGH"
    
    # Major operator signals
    major_ops = {
        "CABOT": "Coterra Energy (formerly Cabot) - major Marcellus operator",
        "COTERRA": "Coterra Energy - major Marcellus operator", 
        "EQT": "EQT Corporation - largest US natural gas producer",
        "SOUTHWESTERN": "SWN - major Appalachian Basin operator",
        "CNX": "CNX Resources - major WV Marcellus operator",
        "ANTERO": "Antero Resources - major Marcellus/Utica operator",
        "EQUINOR": "Equinor - Norwegian major, active in WV",
        "COLUMBIA GAS": "Columbia Gas - major WV pipeline and production",
        "CHESAPEAKE": "Chesapeake Energy - major unconventional operator",
        "ARGUS ENERGY": "Argus Energy - active Lincoln County operator",
    }
    for op_key, op_desc in major_ops.items():
        if op_key in name_up or op_key in desc_up:
            signals.append({"type": "MAJOR_OPERATOR", "weight": 8,
                "note": op_desc, "operator": op_key})
            if priority != "HIGH":
                priority = "HIGH"
    
    # Mineral fraction signals
    if re.search(r"MIN\s", desc_up) or re.search(r"MIN\.", desc_up):
        signals.append({"type": "MINERAL_INTEREST", "weight": 6,
            "note": "Mineral interest (subsurface rights)"})
        if priority == "LOW":
            priority = "MEDIUM"
    
    if "O & G" in desc_up or "OIL" in desc_up and "GAS" in desc_up:
        signals.append({"type": "OIL_GAS_EXPLICIT", "weight": 7,
            "note": "Explicitly described as Oil & Gas mineral rights"})
        if priority == "LOW":
            priority = "MEDIUM"
    
    # Trust/estate signals (classic inherited, forgotten taxes)
    if "TRUSTEE" in name_up or "TRUST" in name_up:
        signals.append({"type": "TRUST_HOLDING", "weight": 3,
            "note": "Trust holding - often forgotten or unmanaged minerals"})
    if " EST" in name_up or "ESTATE" in name_up:
        signals.append({"type": "ESTATE_HOLDING", "weight": 3,
            "note": "Estate holding - heirs may not know about or manage these"})
    if name_up.startswith("CO "):
        signals.append({"type": "CORPORATION", "weight": 4,
            "note": "Corporate entity - check WV SOS for status"})
    
    # Extract acreage
    acre_match = re.search(r"(\d+[\.,]?\d*)\s*(?:AC|ACRE)", desc_up)
    acres = float(acre_match.group(1).replace(",","")) if acre_match else 0
    
    # Extract fraction
    frac_match = re.search(r"(\d+)/(\d+)\s*OF\s*(\d+[\.,]?\d*)\s*AC", desc_up)
    effective_acres = 0
    if frac_match:
        num, den, total = float(frac_match.group(1)), float(frac_match.group(2)), float(frac_match.group(3).replace(",",""))
        effective_acres = (num/den) * total
        signals.append({"type": "FRACTIONAL_INTEREST", "weight": 2,
            "note": f"Fractional mineral interest: {frac_match.group(1)}/{frac_match.group(2)} of {total} acres = {effective_acres:.2f} net mineral acres"})
    elif acres > 0:
        effective_acres = acres
    
    return {
        "signals": signals,
        "priority": priority,
        "acres": acres,
        "effective_acres": effective_acres,
        "priority_score": sum(s["weight"] for s in signals)
    }


def calculate_roi(appraised_value, min_bid_str, county, effective_acres, formation_data):
    """Calculate ROI metrics from all available data."""
    min_bid = float(re.sub(r"[^0-9.]", "", str(min_bid_str))) if min_bid_str else 0
    
    result = {
        "min_bid": min_bid,
        "appraised_value": appraised_value,
        "data_source": "unknown"
    }
    
    county_up = county.upper().replace(" COUNTY","").strip()
    levy = COUNTY_LEVY_RATES.get(county_up, COUNTY_LEVY_RATES["DEFAULT"])
    
    if appraised_value:
        # From actual appraised value
        assessed = appraised_value * 0.60
        actual_tax = (assessed / 100) * levy
        
        # Royalty estimate from appraised value (WV formula: 1.5x-7x)
        royalty_low = appraised_value / 7
        royalty_high = appraised_value / 1.5
        royalty_mid = (royalty_low + royalty_high) / 2
        
        result.update({
            "assessed_value": round(assessed),
            "est_actual_tax": round(actual_tax, 2),
            "royalty_low": round(royalty_low),
            "royalty_high": round(royalty_high),
            "royalty_mid": round(royalty_mid),
            "data_source": "assessor_record"
        })
    elif effective_acres > 0 and formation_data:
        # From formation tier estimates
        avg_per_acre = formation_data.get("avg_royalty_per_acre", 100)
        royalty_est = effective_acres * avg_per_acre
        # Back-calculate appraised value: royalty * 3 (mid-point of 1.5-7x)
        est_appraised = royalty_est * 3
        
        result.update({
            "assessed_value": round(est_appraised * 0.6),
            "est_actual_tax": round((est_appraised * 0.6 / 100) * levy, 2),
            "royalty_low": round(royalty_est * 0.5),
            "royalty_high": round(royalty_est * 2),
            "royalty_mid": round(royalty_est),
            "data_source": "formation_estimate",
            "note": f"Based on {effective_acres:.1f} net mineral acres × ${avg_per_acre}/acre/yr formation average"
        })
    
    # ROI metrics
    if result.get("royalty_mid") and min_bid > 0:
        roy_mid = result["royalty_mid"]
        result["payback_years"] = round(min_bid / roy_mid, 1) if roy_mid > 0 else None
        result["roi_5yr_pct"] = round(((roy_mid * 5 - min_bid) / min_bid) * 100) if min_bid > 0 else None
        result["roi_10yr_pct"] = round(((roy_mid * 10 - min_bid) / min_bid) * 100) if min_bid > 0 else None
        result["roi_rating"] = (
            "EXCEPTIONAL" if result["payback_years"] and result["payback_years"] < 0.5 else
            "EXCELLENT" if result["payback_years"] and result["payback_years"] < 1 else
            "VERY GOOD" if result["payback_years"] and result["payback_years"] < 2 else
            "GOOD" if result["payback_years"] and result["payback_years"] < 5 else
            "MODERATE"
        )
    
    return result


async def scrape_sheriff_async(county, ticket, page):
    """
    Scrape sheriff tax office by ticket number.
    Returns: appraised value, assessed value, actual tax, fee breakdown, book/page.
    The softwaresystems.com platform is used by most WV counties.
    """
    county_low = county.upper().replace(" COUNTY","").strip().lower()
    base = f"http://{county_low}.softwaresystems.com"
    result = {
        "source": "sheriff", "url": base, "success": False,
        "appraised_value": None, "assessed_value": None,
        "actual_tax": None, "penalty": None, "interest": None,
        "publication_fee": None, "total_due": None,
        "tax_years": [],  # breakdown per year
        "book": None, "page": None,
        "owner_address": None,
        "table_rows": [], "raw_text": ""
    }

    try:
        # Navigate directly to results URL - bypass form submission issues
        # Software Systems Inc format: results.html with GET params
        results_url = f"{base}/results.html?TAXYR=&TICKET={ticket}&SUFFIX=&DISTRICT=0&MAP=0000&PARCEL=0000&SUBPARCEL=0000&TAXNAME=&ACCOUNT=&TXTYPE=R&PAIDUNPAID=B&SEARCH=Search"
        print(f"[SHERIFF] Loading results directly: {results_url}", flush=True)
        await page.goto(results_url, timeout=25000, wait_until="domcontentloaded")
        await page.wait_for_timeout(2000)

        body_text = await page.inner_text("body")
        result["raw_text"] = body_text[:3000]
        print(f"[SHERIFF] Results page snippet: {body_text[:400]}", flush=True)

        # Find and click the ticket link in results
        links = await page.query_selector_all("a")
        clicked = False
        for link in links:
            href = (await link.get_attribute("href")) or ""
            txt = (await link.inner_text()).strip()
            if str(ticket) in href or str(ticket) in txt:
                try:
                    await link.click()
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await page.wait_for_timeout(2000)
                    clicked = True
                    print(f"[SHERIFF] Clicked ticket link: {href}", flush=True)
                    break
                except: pass

        # If no link found, try direct ticket URL formats
        if not clicked:
            for yr in ["2025", "2024", "2023", "2022"]:
                ticket_url = f"{base}/ticket.html?TPTYR={yr}&TPTICK={ticket}&TPSX="
                print(f"[SHERIFF] Trying direct ticket URL: {ticket_url}", flush=True)
                await page.goto(ticket_url, timeout=15000, wait_until="domcontentloaded")
                await page.wait_for_timeout(1500)
                body_check = await page.inner_text("body")
                # If we got real data (not just the form), stop here
                if any(kw in body_check.upper() for kw in ["APPRAISED", "ASSESSED", "TAX AMOUNT", "TOTAL DUE"]):
                    print(f"[SHERIFF] Found data at {ticket_url}", flush=True)
                    break
                if str(ticket) in body_check and "OWNER" in body_check.upper():
                    print(f"[SHERIFF] Found ticket detail at year {yr}", flush=True)
                    break

        # Get all table data from detail page
        body_text = await page.inner_text("body")
        result["raw_text"] = body_text[:4000]

        tables = await page.query_selector_all("table")
        all_rows = []
        for table in tables:
            rows_els = await table.query_selector_all("tr")
            for row_el in rows_els:
                cells = await row_el.query_selector_all("td, th")
                texts = [(await c.inner_text()).strip() for c in cells]
                if any(t.strip() for t in texts):
                    all_rows.append(texts)
        result["table_rows"] = all_rows[:50]

        print(f"[SHERIFF] Rows found: {len(all_rows)}", flush=True)
        for r in all_rows[:20]:
            print(f"  ROW: {r}", flush=True)

        # ── PARSE ALL FINANCIAL DATA ──────────────────────────────────────────
        def extract_dollar(text):
            """Extract first dollar amount from text."""
            text = text.replace(",", "").replace("$", "")
            m = re.search(r"\d+\.?\d*", text)
            if m:
                try: return float(m.group())
                except: pass
            return None

        # Parse each row looking for financial fields
        full_text = body_text.upper()
        
        # Appraised and assessed values
        for row in all_rows:
            row_text = " | ".join(row).upper()
            vals = [extract_dollar(c) for c in row if extract_dollar(c) and extract_dollar(c) > 0]
            
            if "APPRAISED" in row_text and not result["appraised_value"]:
                for v in vals:
                    if v > 100:
                        result["appraised_value"] = v
                        result["assessed_value"] = round(v * 0.60)
                        break

            if "ASSESSED" in row_text and vals and not result["assessed_value"]:
                for v in vals:
                    if v > 50:
                        result["assessed_value"] = v
                        break

            # Fee breakdown rows
            if "TAX" in row_text and ("AMOUNT" in row_text or "DUE" in row_text or "CHARGE" in row_text):
                for v in vals:
                    if v > 0 and not result["actual_tax"]:
                        result["actual_tax"] = v

            if "PENALT" in row_text and vals:
                result["penalty"] = vals[0]

            if "INTEREST" in row_text and vals:
                result["interest"] = vals[0]

            if ("PUBLICATION" in row_text or "PUB FEE" in row_text or "ADVERTISING" in row_text) and vals:
                result["publication_fee"] = vals[0]

            if ("TOTAL" in row_text and "DUE" in row_text) and vals:
                result["total_due"] = max(vals)  # largest number in total due row

            # Book and page
            if "BOOK" in row_text and "PAGE" in row_text:
                # Try to extract book and page numbers
                book_m = re.search(r"BOOK[:\s]+([\w\-]+)", row_text)
                page_m = re.search(r"PAGE[:\s]+([\w\-]+)", row_text)
                if book_m: result["book"] = book_m.group(1).strip()
                if page_m: result["page"] = page_m.group(1).strip()
                # Also check adjacent cells
                for i, cell in enumerate(row):
                    if "BOOK" in cell.upper():
                        if i+1 < len(row) and row[i+1].strip():
                            result["book"] = row[i+1].strip()
                    if "PAGE" in cell.upper():
                        if i+1 < len(row) and row[i+1].strip():
                            result["page"] = row[i+1].strip()

            # Tax year breakdown
            year_m = re.search(r"\b(20\d{2}|19\d{2})\b", row_text)
            if year_m and vals and "TAX" in row_text:
                result["tax_years"].append({
                    "year": year_m.group(1),
                    "amount": vals[0] if vals else None,
                    "row": row
                })

        # Also try regex on full body text for appraised value
        if not result["appraised_value"]:
            m = re.search(r"[Aa]ppraised[\s\w]*[:\$]?\s*\$?([\d,]+)", body_text)
            if m:
                try: result["appraised_value"] = float(m.group(1).replace(",",""))
                except: pass

        # Calculate actual tax from assessed value if not found directly
        if result["assessed_value"] and not result["actual_tax"]:
            # WV Class 3 levy ~$0.70 per $100
            result["actual_tax"] = round(result["assessed_value"] * 0.007, 2)

        # Estimate fee breakdown from min_bid if we have actual tax
        # min_bid = actual_tax_years + penalty(5%) + interest(1%/mo) + pub_fee($50-150) + sheriff($25)
        
        if any([result["appraised_value"], result["assessed_value"], result["actual_tax"]]):
            result["success"] = True
            print(f"[SHERIFF] SUCCESS - Appraised: {result['appraised_value']}, Tax: {result['actual_tax']}, Book: {result['book']}/{result['page']}", flush=True)
        else:
            print(f"[SHERIFF] No financial data found in {len(all_rows)} rows", flush=True)

    except Exception as e:
        import traceback
        traceback.print_exc()
        result["error"] = str(e)
        print(f"[SHERIFF] Exception: {e}", flush=True)

    return result



async def run_full_assessment(county, ticket, owner, district, map_num, parcel, min_bid, desc):
    """
    Master assessment engine. Runs sheriff + assessment scrapers in parallel,
    checks production DB, builds fee breakdown, synthesizes with Claude.
    """
    try:
        import anthropic as _anthropic
    except ImportError:
        _anthropic = None
    county_up = county.upper().replace(" COUNTY","").strip()

    # Formation intel (instant, hardcoded)
    formation = FORMATION_DATA.get(county_up, {
        "tier": 3, "marcellus": "UNKNOWN", "utica": "UNKNOWN",
        "desc": f"No formation data for {county_up}",
        "active_operators": [], "avg_royalty_per_acre": 100,
        "drilling_outlook": "UNKNOWN"
    })

    # Description analysis (instant)
    desc_analysis = analyze_description(desc, owner)

    # Check production DB and cached tax data
    production_records = []
    cached_tax = None
    try: production_records = lookup_production_data(owner, county_up)
    except Exception as e: print(f"[ASSESS] Production lookup error: {e}", flush=True)
    if ticket:
        try: cached_tax = get_cached_tax_data(county_up, ticket)
        except: pass

    sheriff_data = {"success": False}
    if cached_tax:
        sheriff_data = {"success": True, "source": "cached",
            "appraised_value": cached_tax.get("appraised_value"),
            "assessed_value": cached_tax.get("assessed_value"),
            "actual_tax": cached_tax.get("actual_tax"),
            "penalty": cached_tax.get("penalty"),
            "interest": cached_tax.get("interest"),
            "publication_fee": cached_tax.get("publication_fee"),
            "book": cached_tax.get("book"), "page": cached_tax.get("page")}
        print(f"[ASSESS] Using cached tax data for ticket {ticket}", flush=True)

    # Run scrapers in parallel
    assessment_data = {"success": False}
    well_data_result = {"success": False, "wells": []}

    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
        ctx = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            ignore_https_errors=True)

        pg2 = await ctx.new_page()
        pg3 = await ctx.new_page()

        if not cached_tax and ticket:
            pg1 = await ctx.new_page()
            results = await asyncio.gather(
                scrape_sheriff_async(county_up, ticket, pg1),
                return_exceptions=True)
            sheriff_data = results[0] if not isinstance(results[0], Exception) else {"success": False, "error": str(results[0])}
        # Also scrape WVDEP wells (separate from browser context)
        try:
            well_list = await scrape_wvdep_wells(county_up.title())
            well_data_result = {"success": bool(well_list), "wells": [{"text": str(w)} for w in well_list[:10]]}
        except Exception as e:
            well_data_result = {"success": False, "wells": [], "error": str(e)}

        await browser.close()

    # Store new sheriff data
    if sheriff_data.get("success") and ticket and not cached_tax:
        try: store_og_tax_data(county_up, ticket, sheriff_data)
        except: pass

    print(f"[ASSESS] Sheriff:{sheriff_data.get('success')} Assess:{assessment_data.get('success')} Wells:{well_data_result.get('success')} Prod:{len(production_records)}", flush=True)

    # Best value from sources
    appraised_value = sheriff_data.get("appraised_value") or assessment_data.get("appraised_value")
    value_source = "sheriff" if sheriff_data.get("appraised_value") else "assessment" if assessment_data.get("appraised_value") else "none"

    # ROI
    roi = calculate_roi(appraised_value, min_bid, county_up, desc_analysis.get("effective_acres",0), formation)

    # Fee breakdown
    actual_tax = sheriff_data.get("actual_tax")
    fee_breakdown = {}
    if actual_tax and roi.get("min_bid",0) > 0:
        mb = roi["min_bid"]
        fees = max(0, mb - actual_tax)
        fee_breakdown = {
            "actual_tax": actual_tax,
            "fees_and_penalties": round(fees, 2),
            "tax_pct_of_bid": round((actual_tax/mb)*100, 1),
            "fee_pct_of_bid": round((fees/mb)*100, 1),
            "penalty": sheriff_data.get("penalty"),
            "interest": sheriff_data.get("interest"),
            "publication_fee": sheriff_data.get("publication_fee"),
        }

    book = sheriff_data.get("book")
    page_num = sheriff_data.get("page")
    idx_url = f"https://www.courtplus.com/cgi-bin/docdetail.cgi?county={county_up.lower()}&book={book}&page={page_num}" if book and page_num else None

    confirmed_producing = any(float(r.get("gas_mcf",0) or 0) > 0 or float(r.get("oil_bbl",0) or 0) > 0 for r in production_records)

    prod_lines = "\n".join([f"  {r.get('year')}: Gas={r.get('gas_mcf','?')} MCF Oil={r.get('oil_bbl','?')} BBL" for r in production_records[:5]]) or "  Not found in WVDEP database"
    wells_lines = "\n".join([w.get("text","")[:100] for w in well_data_result.get("wells",[])[:6]]) or "No well data"

    prompt = f"""You are an expert WV oil and gas mineral rights investment analyst.

PROPERTY: {owner} | {county} | District: {district}
DESCRIPTION: {desc}
MIN BID: {min_bid}  TICKET: {ticket}

FEE BREAKDOWN:
  Actual property tax: ${actual_tax or "not retrieved"}
  Publication/penalty fees: ${fee_breakdown.get("fees_and_penalties","?") if fee_breakdown else "not calculated"}
  Tax is {fee_breakdown.get("tax_pct_of_bid","?")}% of bid | Fees are {fee_breakdown.get("fee_pct_of_bid","?")}%

ASSESSED VALUES (source: {value_source}):
  Appraised: ${appraised_value or "not retrieved"}
  Assessed (60%): ${roi.get("assessed_value","?")}
  Book/Page: {book or "not found"}/{page_num or "not found"}

WVDEP PRODUCTION FOR THIS OWNER:
{prod_lines}
{"*** CONFIRMED PRODUCING - active royalty income! ***" if confirmed_producing else ""}

FORMATION: {county_up} Tier {formation.get("tier")} - {formation.get("marcellus","?")} Marcellus
{formation.get("desc","")}

WELLS IN DISTRICT:
{wells_lines}

EST ROI: Bid ${roi.get("min_bid",0):,.2f} | Royalty est ${roi.get("royalty_low",0):,.0f}-${roi.get("royalty_high",0):,.0f}/yr | Payback {roi.get("payback_years","?")} yrs | 10yr ROI {roi.get("roi_10yr_pct","?")}%

Provide:
1. INVESTMENT GRADE (A+ to F)
2. IS IT PRODUCING? (use production DB evidence - be specific about MCF numbers)
3. FEE BREAKDOWN - what % of bid is real tax vs fees?
4. WHAT YOU'RE BUYING - plain English
5. VALUATION - appraised vs what you pay
6. DRILLING POTENTIAL - formation + operators
7. RECOMMENDATION: BID / SKIP / INVESTIGATE
8. ACTION: pull deed book {book}/{page_num} to verify mineral reservation language"""

    try:
        if not _anthropic:
            claude_assessment = "AI assessment unavailable: anthropic package not installed on server"
        else:
            client = _anthropic.Anthropic()
            msg = client.messages.create(model="claude-opus-4-6", max_tokens=1500,
                messages=[{"role":"user","content":prompt}])
            claude_assessment = msg.content[0].text
    except Exception as e:
        claude_assessment = f"AI assessment unavailable: {e}"

    return {
        "success": True, "county": county_up, "ticket": ticket, "owner": owner, "min_bid": min_bid,
        "priority": desc_analysis.get("priority","LOW"),
        "priority_score": desc_analysis.get("priority_score",0),
        "signals": desc_analysis.get("signals",[]),
        "formation": formation, "roi": roi,
        "fee_breakdown": fee_breakdown,
        "book": book, "page": page_num, "idx_url": idx_url,
        "confirmed_producing": confirmed_producing,
        "production_records": production_records,
        "sources": {
            "sheriff": {"success": sheriff_data.get("success"), "appraised": sheriff_data.get("appraised_value"), "source": sheriff_data.get("source","scraped")},
            "assessment": {"success": assessment_data.get("success")},
            "wells": {"success": well_data_result.get("success"), "count": len(well_data_result.get("wells",[]))}
        },
        "raw_data": {"sheriff": sheriff_data},
        "assessment": claude_assessment
    }

def run_assessment(county, ticket, owner, district, map_num, parcel, min_bid, desc):
    """Synchronous entry point."""
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        result = loop.run_until_complete(
            run_full_assessment(county, ticket, owner, district, map_num, parcel, min_bid, desc)
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        result = {"success": False, "error": str(e)}
    finally:
        loop.close()
    return result




# ═══════════════════════════════════════════════════════════════════
# O&G DATA BANK - Download WVDEP production data into Supabase
# ═══════════════════════════════════════════════════════════════════
SUPABASE_URL_DB = "https://uhunhyfgwvoknqnkzlmr.supabase.co"
SUPABASE_KEY_DB = "sb_publishable_X1nUMQ4GQfiPj-AsVvigwQ_7g3d4i95"
BANK_BUILD_STATUS = {"state": "idle", "progress": "", "records": 0, "errors": []}

WVDEP_PRODUCTION_URLS = {
    2024: "https://apps.dep.wv.gov/Documents/OOG/ProductionReports/2020-2029/2024Production.xlsx",
    2023: "https://apps.dep.wv.gov/Documents/OOG/ProductionReports/2020-2029/2023Production.xlsx",
    2022: "https://apps.dep.wv.gov/Documents/OOG/ProductionReports/2020-2029/2022Production.xlsx",
}


async def build_production_data_bank():
    """Download WVDEP annual production Excel files and load into Supabase."""
    import openpyxl, io, json, urllib.request, urllib.error
    global BANK_BUILD_STATUS
    BANK_BUILD_STATUS = {"state": "running", "progress": "Starting...", "records": 0, "errors": []}
    total = 0

    for year, url in WVDEP_PRODUCTION_URLS.items():
        try:
            BANK_BUILD_STATUS["progress"] = f"Downloading {year}..."
            print(f"[BANK] Downloading {year}: {url}", flush=True)
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=120) as r:
                excel_bytes = r.read()
            print(f"[BANK] {year} downloaded: {len(excel_bytes)} bytes", flush=True)

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip().upper() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            print(f"[BANK] {year} headers: {headers[:12]}", flush=True)

            def col_idx(names):
                for n in names:
                    for i, h in enumerate(headers):
                        if n.upper() in h: return i
                return None

            api_i  = col_idx(["API"])
            cty_i  = col_idx(["COUNTY"])
            own_i  = col_idx(["OWNER","OPERATOR","COMPANY","LESSEE"])
            gas_i  = col_idx(["GAS","MCF"])
            oil_i  = col_idx(["OIL","BBL"])
            dst_i  = col_idx(["DISTRICT","DIST"])

            batch = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                def sv(i): return str(row[i] or "").strip() if i is not None and i < len(row) else ""
                def nv(i):
                    try: return float(row[i] or 0) if i is not None and i < len(row) else 0
                    except: return 0
                rec = {
                    "api_number": sv(api_i), "county": sv(cty_i).upper(),
                    "district": sv(dst_i).upper(), "owner_name": sv(own_i).upper(),
                    "year": year, "gas_mcf": nv(gas_i), "oil_bbl": nv(oil_i),
                    "operator": sv(own_i).upper()
                }
                if rec["county"] or rec["api_number"]:
                    batch.append(rec)
                if len(batch) >= 200:
                    _supabase_insert("og_production", batch)
                    total += len(batch)
                    BANK_BUILD_STATUS["records"] = total
                    BANK_BUILD_STATUS["progress"] = f"{year}: {total} records"
                    batch = []
            if batch:
                _supabase_insert("og_production", batch)
                total += len(batch)

            print(f"[BANK] {year} done", flush=True)

        except Exception as e:
            import traceback; traceback.print_exc()
            BANK_BUILD_STATUS["errors"].append(f"{year}: {e}")

    BANK_BUILD_STATUS["state"] = "complete"
    BANK_BUILD_STATUS["progress"] = f"Done. {total} records loaded."
    BANK_BUILD_STATUS["records"] = total
    print(f"[BANK] Complete: {total} records", flush=True)


def _supabase_insert(table, batch):
    """Insert batch into Supabase table."""
    import json, urllib.request, urllib.error
    data = json.dumps(batch).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL_DB}/rest/v1/{table}",
        data=data,
        headers={"apikey": SUPABASE_KEY_DB, "Authorization": f"Bearer {SUPABASE_KEY_DB}",
                 "Content-Type": "application/json", "Prefer": "resolution=ignore-duplicates"},
        method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=30): pass
    except urllib.error.HTTPError as e:
        print(f"[DB] Insert {table} error: {e.code} {e.read()[:300]}", flush=True)


def lookup_production_data(owner_name, county):
    """Look up production data for owner from Supabase."""
    import json, urllib.request, urllib.parse
    name_parts = owner_name.upper().strip().split()
    search = name_parts[0] if name_parts else owner_name.upper()
    county_up = county.upper().replace(" COUNTY","").strip()
    try:
        qs = f"owner_name=ilike.*{urllib.parse.quote(search)}*&county=eq.{urllib.parse.quote(county_up)}&order=year.desc&limit=10"
        req = urllib.request.Request(
            f"{SUPABASE_URL_DB}/rest/v1/og_production?{qs}",
            headers={"apikey": SUPABASE_KEY_DB, "Authorization": f"Bearer {SUPABASE_KEY_DB}"}
        )
        with urllib.request.urlopen(req, timeout=8) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f"[LOOKUP] {e}", flush=True)
        return []


def store_og_tax_data(county, ticket, data):
    """Store sheriff scrape result in og_tax_data table."""
    import json
    record = {
        "county": county.upper().replace(" COUNTY","").strip(), "ticket": str(ticket),
        "appraised_value": data.get("appraised_value"), "assessed_value": data.get("assessed_value"),
        "actual_tax": data.get("actual_tax"), "penalty": data.get("penalty"),
        "interest": data.get("interest"), "publication_fee": data.get("publication_fee"),
        "total_due": data.get("total_due"), "book": data.get("book"), "page": data.get("page"),
    }
    _supabase_insert("og_tax_data", [record])


def get_cached_tax_data(county, ticket):
    """Check if we already have sheriff data for this ticket."""
    import json, urllib.request, urllib.parse
    county_up = county.upper().replace(" COUNTY","").strip()
    try:
        qs = f"county=eq.{urllib.parse.quote(county_up)}&ticket=eq.{urllib.parse.quote(str(ticket))}&limit=1"
        req = urllib.request.Request(
            f"{SUPABASE_URL_DB}/rest/v1/og_tax_data?{qs}",
            headers={"apikey": SUPABASE_KEY_DB, "Authorization": f"Bearer {SUPABASE_KEY_DB}"}
        )
        with urllib.request.urlopen(req, timeout=8) as r:
            rows = json.loads(r.read())
            return rows[0] if rows else None
    except:
        return None

# ═══════════════════════════════════════════════════════════════════
class Handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args): pass

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Content-Type')

    def do_GET(self):
        path = self.path.split("?")[0]
        print(f"[GET] {path}", flush=True)
        if path == "/counties":
            return self.respond({"success": True, "counties": get_county_registry()})

        if path == "/proxy":
            from urllib.parse import parse_qs, urlparse
            qs = parse_qs(urlparse(self.path).query)
            target = qs.get('url', [None])[0]
            if not target:
                return self.respond({"error": "No URL provided"})
            try:
                import urllib.request as ur
                req = ur.Request(target, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8',
                })
                with ur.urlopen(req, timeout=15) as r:
                    html = r.read().decode('utf-8', errors='replace')
                return self.respond({"contents": html})
            except Exception as e:
                return self.respond({"error": str(e)})

        if path == "/wvsao-sync":
            return self.respond(sync_wvsao_dates())

        if path == "/sheriff-lookup":
            from urllib.parse import parse_qs, urlparse
            qs = parse_qs(urlparse(self.path).query)
            get = lambda k: qs.get(k, [''])[0]
            county  = get('county')
            ticket  = get('ticket')
            tax_year = get('year') or None
            if not county or not ticket:
                return self.respond({"error": "county and ticket required"})
            result = run_sheriff_lookup(county, ticket, tax_year=tax_year)
            return self.respond(result)


        if path == "/build-data-bank":
            """Download WVDEP production data and load into Supabase."""
            import threading
            def run_build():
                asyncio.run(build_production_data_bank())
            t = threading.Thread(target=run_build, daemon=True)
            t.start()
            return self.respond({"status": "started", "message": "Building O&G data bank in background. Check /bank-status for progress."})

        if path == "/bank-status":
            return self.respond({"status": BANK_BUILD_STATUS})

        if path == "/og-intel":
            from urllib.parse import parse_qs, urlparse
            qs = parse_qs(urlparse(self.path).query)
            get = lambda k: qs.get(k, [''])[0]
            owner   = get('owner')
            county  = get('county')
            district= get('district')
            map_num = get('map')
            parcel  = get('parcel')
            min_bid = get('minBid')
            desc    = get('desc')
            if not owner or not county:
                return self.respond({"error": "owner and county required"})
            try:
                result = run_assessment(county, get('ticket') or '0', owner, district, map_num, parcel, min_bid, desc)
                return self.respond(result)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return self.respond({"error": str(e)})

        if path == "/og-assess":
            from urllib.parse import parse_qs, urlparse
            qs = parse_qs(urlparse(self.path).query)
            get = lambda k: qs.get(k, [''])[0]
            county  = get('county')
            ticket  = get('ticket')
            owner   = get('owner')
            district= get('district')
            map_num = get('map')
            parcel  = get('parcel')
            min_bid = get('minBid')
            desc    = get('desc')
            if not county or not owner:
                return self.respond({"error": "county and owner required"})
            try:
                result = run_assessment(county, ticket, owner, district, map_num, parcel, min_bid, desc)
                return self.respond(result)
            except Exception as e:
                import traceback
                traceback.print_exc()
                return self.respond({"error": str(e)})

        body = b'WV Tax Lien API - PDF + CAMA (55 counties) + IDX'
        self.send_response(200)
        self.send_header('Content-Type','text/plain')
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Content-Length',len(body))
        self.end_headers(); self.wfile.write(body)

    def do_POST(self):
        try:
            ct = self.headers.get('Content-Type','')
            length = int(self.headers.get('Content-Length',0))
            body = self.rfile.read(length)
            path = self.path.split("?")[0]

            if path == "/cama":
                data = json.loads(body)
                return self.respond(lookup_cama(
                    data.get("county",""), data.get("dist","01"),
                    data.get("map","0001"), data.get("parcel","0001")))

            if path == "/idx":
                data = json.loads(body)
                return self.respond(search_idx(data.get("county",""), data.get("name","")))

            if path == "/mapwv":
                data = json.loads(body)
                return self.respond(fetch_mapwv_owner(
                    data.get("url",""),
                    data.get("countyKey",""),
                    data.get("map",""),
                    data.get("parcel","")
                ))

            if path == "/assessment":
                data = json.loads(body)
                return self.respond(fetch_assessment_detail(data.get("pid","")))

            if path == "/idx-search":
                data = json.loads(body)
                return self.respond(do_title_search(
                    data.get("county",""),
                    data.get("deed_book",""),
                    data.get("deed_page",""),
                    data.get("owner_name",""),
                    int(data.get("years_back", 25))
                ))

            if path == "/analyze":
                data = json.loads(body)
                return self.respond(call_claude(data.get("prompt","")))

            if path == "/idx-screenshot":
                try:
                    import base64
                    with open("/tmp/idx_screenshot.png","rb") as f:
                        img = base64.b64encode(f.read()).decode()
                    return self.respond({"success":True,"image":img})
                except Exception as e:
                    return self.respond({"success":False,"error":str(e)})

            if path == "/og-intel":
                import asyncio
                body = json.loads(self.rfile.read(int(self.headers['Content-Length'])))
                county = body.get('county','').upper().replace(' COUNTY','').strip()
                district = body.get('district','')
                owner = body.get('owner','')
                description = body.get('description','')
                min_bid = body.get('minBid','')
                print(f"[OG-INTEL] county={county} district={district} owner={owner}", flush=True)

                # Scrape WVDEP for active H6A wells in this county
                try:
                    well_data = asyncio.run(scrape_wvdep_wells(county.title()))
                    print(f"[OG-INTEL] Found {len(well_data)} wells", flush=True)
                except Exception as e:
                    print(f"[OG-INTEL] Scrape error: {e}", flush=True)
                    well_data = []

                # Build assessment
                assessment = og_intel_assessment(county, district, owner, description, min_bid, well_data)
                return self.respond({"success": True, "assessment": assessment, "well_count": len(well_data)})

            if 'multipart/form-data' not in ct:
                return self.respond({'success':False,'error':'Expected multipart/form-data'})

            boundary = ct.split('boundary=')[1].strip().encode()
            pdf_data = None
            for part in body.split(b'--' + boundary):
                if b'filename=' in part and b'.pdf' in part.lower():
                    hend = part.find(b'\r\n\r\n')
                    if hend != -1:
                        pdf_data = part[hend+4:].rstrip(b'\r\n-')
                        break

            if not pdf_data:
                return self.respond({'success':False,'error':'No PDF found'})
            self.respond(parse_pdf(pdf_data))

        except Exception as e:
            self.respond({'success':False,'error':str(e)})

    def respond(self, data):
        body = json.dumps(data).encode()
        self.send_response(200)
        self.send_header('Content-Type','application/json')
        self._cors()
        self.send_header('Content-Length',len(body))
        self.end_headers(); self.wfile.write(body)


# ── PDF PARSING ───────────────────────────────────────────────────────────────

def extract_county_from_line(line):
    upper = line.upper().strip()
    m = re.match(r'^([A-Z]+(?:\s+[A-Z]+)?)\s+COUNTY$', upper)
    if m and m.group(1).strip() in WV_COUNTIES:
        return f"{m.group(1).strip()} COUNTY"
    return None

def parse_pdf(pdf_bytes):
    result = {'county':'','date':'','time':'','location':'','rows':[],
              'lastUpdated':datetime.now(timezone.utc).isoformat()}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            lines = [l.strip() for l in (pdf.pages[0].extract_text() or '').split('\n') if l.strip()]
            county_found = date_found = False
            for i, line in enumerate(lines[:20]):
                if any(s in line for s in ['WEB HANDOUT','CERTIFICATE','TICKET','DISTRICT','ASSESSED','LEGAL','MINIMUM']): continue
                if re.match(r'^\d{4}-C-\d+', line): break
                if not county_found:
                    county = extract_county_from_line(line)
                    if county: result['county'] = county; county_found = True; continue
                if not date_found and '/' in line and len(line) < 35:
                    dm = re.match(r'^(\d{1,2}/\d{1,2}/\d{4})\s+(\d{1,2}:\d{2}\s*[AP]M)$', line, re.I)
                    if dm:
                        result['date'] = dm.group(1); result['time'] = dm.group(2).strip(); date_found = True
                        for j in range(i+1, min(i+5, len(lines))):
                            nl = lines[j]
                            if not any(s in nl for s in ['CERTIFICATE','TICKET','DISTRICT','ASSESSED']):
                                if not re.match(r'^\d{4}-C-\d+', nl):
                                    result['location'] = nl; break
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    for row in table:
                        if not row or not row[0]: continue
                        cert = (row[0] or '').replace('\n',' ').strip()
                        if not re.match(r'^\d{4}-C-\d+$', cert): continue
                        result['rows'].append({
                            'cert': cert,
                            'ticket': (row[1] or '').replace('\n',' ').strip(),
                            'district': (row[2] or '').replace('\n',' ').strip(),
                            'map': (row[3] or '').replace('\n',' ').strip(),
                            'parcel': (row[4] or '').replace('\n',' ').strip(),
                            'sub': (row[5] or '0000').replace('\n',' ').strip() or '0000',
                            'subsub': (row[6] or '0000').replace('\n',' ').strip() or '0000',
                            'name': (row[7] or '').replace('\n',' ').strip(),
                            'desc': (row[8] or '').replace('\n',' ').strip(),
                            'minBid': (row[9] or '').replace('\n',' ').strip(),
                        })
        if not result['county']: return {'success':False,'error':'Could not identify WV county.'}
        if not result['rows']: return {'success':False,'error':'No property rows found.'}
        return {'success':True,'data':result}
    except Exception as e:
        return {'success':False,'error':f'PDF parsing error: {str(e)}'}


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    print(f'WV Tax Lien API running on port {port} — 55 counties CAMA enabled')
    ensure_chromium()
    HTTPServer(('0.0.0.0', port), Handler).serve_forever()


# ── O&G INTEL - PLAYWRIGHT SCRAPER ───────────────────────────────────────────
# Scrapes WVDEP well database for active H6A (Marcellus/Utica) wells by county
# Cross-references with district/corp to score tax lien O&G potential

# Formation tiers by county - based on WVGES 2022 production data
OG_FORMATION_TIERS = {
    # Tier 1 - Top Marcellus AND Utica producers
    'MARSHALL':{'marcellus':1,'utica':1,'notes':'#1 county both formations, wet gas window'},
    'WETZEL':{'marcellus':1,'utica':2,'notes':'Top Marcellus producer, Southwestern Energy hub'},
    'TYLER':{'marcellus':1,'utica':1,'notes':'Top Marcellus liquids, active Utica drilling'},
    'DODDRIDGE':{'marcellus':1,'utica':2,'notes':'Strong Marcellus, heavy drilling activity'},
    'RITCHIE':{'marcellus':1,'utica':2,'notes':'Prolific conventional + Marcellus'},
    'PLEASANTS':{'marcellus':1,'utica':2,'notes':'Active Marcellus drilling corridor'},
    'BROOKE':{'marcellus':1,'utica':2,'notes':'Northern panhandle wet gas'},
    'OHIO':{'marcellus':1,'utica':2,'notes':'Northern panhandle, highest royalty rates'},
    # Tier 2 - Strong production
    'WIRT':{'marcellus':2,'utica':2,'notes':'Active Marcellus corridor'},
    'WOOD':{'marcellus':2,'utica':2,'notes':'Parkersburg area, pipeline infrastructure'},
    'JACKSON':{'marcellus':2,'utica':3,'notes':'Moderate Marcellus activity'},
    'ROANE':{'marcellus':2,'utica':3,'notes':'Conventional + Marcellus mix'},
    'CALHOUN':{'marcellus':2,'utica':3,'notes':'Some Marcellus, mostly conventional'},
    'GILMER':{'marcellus':2,'utica':3,'notes':'Conventional O&G, some Marcellus'},
    'KANAWHA':{'marcellus':2,'utica':3,'notes':'Large county, active in northern districts'},
    'PUTNAM':{'marcellus':2,'utica':3,'notes':'Moderate activity, near Kanawha hub'},
    'LINCOLN':{'marcellus':2,'utica':3,'notes':'Some Marcellus, active conventional'},
    'WAYNE':{'marcellus':2,'utica':3,'notes':'Southern activity corridor'},
    'MINGO':{'marcellus':2,'utica':3,'notes':'CBM and conventional'},
    'LOGAN':{'marcellus':2,'utica':3,'notes':'CBM heavy, some Marcellus'},
    'BRAXTON':{'marcellus':2,'utica':3,'notes':'Moderate conventional and Marcellus'},
    'NICHOLAS':{'marcellus':2,'utica':3,'notes':'Some Marcellus in northern districts'},
    'WEBSTER':{'marcellus':3,'utica':3,'notes':'Limited Marcellus'},
    'CLAY':{'marcellus':3,'utica':3,'notes':'Limited activity'},
}

async def scrape_wvdep_wells(county, operator='', status='Active Well', permit_type='Horizontal 6A Well'):
    """Use Playwright to scrape WVDEP well database for a county."""
    try:
        from playwright.async_api import async_playwright
        import asyncio

        results = []
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            page = await browser.new_page()
            await page.goto('https://tagis.dep.wv.gov/oog/', timeout=30000)
            await page.wait_for_load_state('networkidle', timeout=15000)

            # Select county
            await page.select_option('select[name*="county"], select[id*="county"], select', 
                                     label=county.title(), timeout=5000)

            # Select well status
            if status:
                try:
                    await page.select_option('select[name*="status"], select[id*="status"]',
                                             label=status, timeout=3000)
                except: pass

            # Select permit type (H6A = Marcellus/Utica horizontal)
            if permit_type:
                try:
                    await page.select_option('select[name*="permit"], select[id*="permit"]',
                                             label=permit_type, timeout=3000)
                except: pass

            # Set operator if provided
            if operator:
                try:
                    await page.fill('input[name*="operator"], input[id*="operator"]', operator)
                except: pass

            # Click search
            await page.click('input[type="submit"], button[type="submit"]', timeout=5000)
            await page.wait_for_load_state('networkidle', timeout=20000)
            await asyncio.sleep(2)

            # Parse results table
            html = await page.content()
            await browser.close()

            # Parse the results table
            rows = re.findall(
                r'<tr[^>]*>(.*?)</tr>', html, re.S | re.I
            )
            for row in rows[1:]:  # skip header
                cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.S | re.I)
                cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
                if len(cells) >= 8:
                    results.append({
                        'permit_id': cells[1] if len(cells)>1 else '',
                        'permit_type': cells[3] if len(cells)>3 else '',
                        'issued': cells[4] if len(cells)>4 else '',
                        'operator': cells[5] if len(cells)>5 else '',
                        'status': cells[6] if len(cells)>6 else '',
                        'well_type': cells[7] if len(cells)>7 else '',
                        'well_use': cells[8] if len(cells)>8 else '',
                        'formation': cells[12] if len(cells)>12 else '',
                        'lat': cells[10] if len(cells)>10 else '',
                        'lon': cells[11] if len(cells)>11 else '',
                    })

        return results

    except Exception as e:
        import traceback
        traceback.print_exc()
        return []


def og_intel_assessment(county, district, owner_name, description, min_bid, well_data):
    """Build O&G intelligence assessment from all available data."""
    county_up = county.upper().replace(' COUNTY','').strip()
    district_up = district.upper().strip() if district else ''

    # Formation tier
    tier_info = OG_FORMATION_TIERS.get(county_up, {
        'marcellus':3,'utica':3,'notes':'Limited formation data available'
    })
    marc_tier = tier_info['marcellus']
    utica_tier = tier_info['utica']

    # Corp limit flag - corp districts are incorporated towns
    # O&G companies need rights within radius of unit - corp limits = infrastructure exists
    is_corp = 'CORP' in district_up or 'CORPORATION' in district_up
    corp_note = ''
    if is_corp:
        town = district_up.replace('CORP','').replace('CORPORATION','').strip().title()
        corp_note = (f"Property is within {town} corporate limits. "
                    f"Corp limit parcels often sit within or adjacent to active drilling units — "
                    f"O&G companies must acquire rights within ~1,500ft radius of horizontal bore. "
                    f"Infrastructure (roads, pipelines) likely already in place.")

    # Parse description for mineral indicators
    desc_up = description.upper() if description else ''
    is_mineral = any(k in desc_up for k in ['MIN ', 'MINERAL', 'O & G', 'O&G', 'GAS', 'OIL',
                                              'ROYALT', 'WORKING INT', 'WI ', '1/8', '1/6',
                                              '1/4', '1/16', 'MCF', 'BBL'])
    mineral_fraction = ''
    frac_match = re.search(r'((?:\d+/\d+\s+OF\s+)+[\d\.,]+ ?AC)', desc_up)
    if frac_match:
        mineral_fraction = frac_match.group(1)

    # Acreage from description
    acres_match = re.search(r'([\d\.]+)\s*AC', desc_up)
    acres = float(acres_match.group(1)) if acres_match else None

    # Well data analysis
    active_wells = [w for w in well_data if 'active' in w.get('status','').lower()]
    h6a_wells = [w for w in well_data if 'horizontal 6a' in w.get('permit_type','').lower() or 
                 'h6a' in w.get('permit_type','').lower()]
    operators = list(set(w['operator'] for w in active_wells if w.get('operator')))
    formations = list(set(w['formation'] for w in well_data if w.get('formation') and 
                          w['formation'].strip() not in ['','N/A','Not Available']))

    # Score calculation
    score = 0
    signals = []

    # Formation tier scoring
    if marc_tier == 1:
        score += 40
        signals.append(f"🔥 Top-tier Marcellus county ({county_up})")
    elif marc_tier == 2:
        score += 25
        signals.append(f"🟡 Active Marcellus county ({county_up})")
    else:
        score += 5
        signals.append(f"⚪ Limited Marcellus activity in {county_up}")

    if utica_tier == 1:
        score += 20
        signals.append("🔥 Prime Utica/Point Pleasant zone")
    elif utica_tier == 2:
        score += 10
        signals.append("🟡 Utica potential present")

    # Active H6A wells in county
    if len(h6a_wells) > 50:
        score += 25
        signals.append(f"🔥 {len(h6a_wells)} active H6A horizontal wells in county")
    elif len(h6a_wells) > 10:
        score += 15
        signals.append(f"🟡 {len(h6a_wells)} H6A horizontal wells in county")
    elif len(h6a_wells) > 0:
        score += 8
        signals.append(f"⚪ {len(h6a_wells)} H6A wells in county")

    # Corp limit bonus
    if is_corp:
        score += 15
        signals.append(f"🏘️ Corp limit property — unit radius likely includes this parcel")

    # Mineral description bonus
    if is_mineral:
        score += 15
        signals.append("⛏️ Mineral/O&G interest confirmed in legal description")
    if mineral_fraction:
        score += 5
        signals.append(f"📐 Fractional interest: {mineral_fraction}")

    # Operator signals
    major_operators = ['EQT','CNX','SOUTHWESTERN','SWN','ANTERO','TARGA','CHESAPEAKE',
                       'CHEVRON','COLUMBIA','CABOT','RANGE RESOURCES','DOMINION',
                       'EQUINOR','DIVERSIFIED','CARDINAL MIDSTREAM','HALL DRILLING']
    found_majors = [op for op in operators for maj in major_operators 
                    if maj in op.upper()]
    if found_majors:
        score += 20
        signals.append(f"🏢 Major operators active in county: {', '.join(set(found_majors[:3]))}")

    # Min bid vs potential signal
    try:
        bid = float(str(min_bid).replace('$','').replace(',',''))
        if bid < 300 and is_mineral and marc_tier <= 2:
            score += 10
            signals.append(f"💰 Very low min bid (${bid:.2f}) for mineral interest in active formation county")
    except: pass

    # Score to rating
    if score >= 80:
        rating = "🔥 HIGH PRIORITY"
        summary = "Strong indicators of active O&G production or imminent drilling unit inclusion."
    elif score >= 50:
        rating = "🟡 MODERATE POTENTIAL"
        summary = "Formation present and some activity. Worth investigating further."
    elif score >= 25:
        rating = "⚪ LOW-MODERATE"
        summary = "Some O&G potential but limited active indicators for this specific property."
    else:
        rating = "⬜ LOW"
        summary = "Limited O&G signals. County not in primary formation zone."

    return {
        'rating': rating,
        'score': score,
        'summary': summary,
        'signals': signals,
        'corp_note': corp_note,
        'is_corp': is_corp,
        'is_mineral': is_mineral,
        'formation_tier': f"Marcellus T{marc_tier} / Utica T{utica_tier}",
        'formation_notes': tier_info['notes'],
        'active_wells_in_county': len(active_wells),
        'h6a_wells_in_county': len(h6a_wells),
        'operators': operators[:5],
        'formations_found': formations[:5],
        'mineral_fraction': mineral_fraction,
        'acres': acres,
    }
